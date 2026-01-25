"""
Backend - Vaucher e Álvares Sistema de Cadastro
FastAPI + PostgreSQL + Geração de Documentos + Resend para E-mail
Com gerenciamento de usuários no banco de dados
VERSÃO 3.0 - COM PORTAL DO CLIENTE
"""

from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Depends, Header, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, EmailStr
from typing import Optional, List
import os
import json
import zipfile
import shutil
from datetime import datetime, timezone

# Rate Limiting
from slowapi import _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded

import uuid
import httpx
import base64
import secrets
from io import BytesIO
from dateutil.relativedelta import relativedelta

# Configurações
from modules.config import (
    BASE_DIR,
    MODELOS_DIR,
    UPLOADS_DIR,
    GERADOS_DIR,
    STATIC_DIR,
    RESEND_API_KEY,
    FROM_EMAIL,
    ALLOWED_ORIGINS,
    logger,
    limiter,
)

# Funções de segurança
from modules.security import (
    decodificar_token_cliente,
    criar_email_html,
    validar_arquivo,
    sanitizar_nome_arquivo,
)

# Funções de banco de dados
from modules.database import (
    get_db,
    init_db,
    salvar_cadastro,
    carregar_cadastros,
    buscar_cadastro,
    atualizar_status,
    salvar_financeiro,
    buscar_financeiro,
    atualizar_status_prestacao,
    # Documentos (usados em /api/cliente/documentos)
    listar_documentos_admin,
    buscar_documento_admin,
    listar_documentos_extras,
    buscar_documento_extra,
    # Auth cliente (usado em verificar_token_cliente)
    buscar_cliente_auth,
)

# Modelos Pydantic
from modules.models import (
    DadosCliente,
    SalvarRascunhoDemanda,
    FinanceiroData,
    SolicitacaoAtualizacao,
    EnvioAtualizacao,
    RejeicaoAtualizacao,
)

# Gerador de documentos
from modules.documents import gerador, gerar_peticao_auxilio_moradia

# Funções de e-mail
from modules.email import enviar_email_resend, enviar_email_assinatura_digital

# Assinatura digital
from modules.assinatura import (
    gerar_link_govbr,
    criar_documento_zapsign,
    verificar_status_documento,
    obter_documento_assinado,
)

# PostgreSQL
import psycopg2
from psycopg2.extras import RealDictCursor

# ============================================
# CONFIGURAÇÃO DA APLICAÇÃO
# ============================================

app = FastAPI(
    title="Vaucher e Álvares - API",
    description="Sistema de cadastro de clientes e geração de documentos",
    version="3.1.0"  # Atualizado com melhorias de segurança
)

# Rate Limiter - proteção contra brute force e DDoS (usando instância compartilhada)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# CORS - permitir acesso dos frontends (usando ALLOWED_ORIGINS do config.py)
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "Accept"],
)

# Criar diretórios se não existirem
for dir_path in [MODELOS_DIR, UPLOADS_DIR, GERADOS_DIR, STATIC_DIR]:
    os.makedirs(dir_path, exist_ok=True)

# Servir arquivos estáticos (logo)
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# ============================================
# VERIFICAÇÃO DE AUTENTICAÇÃO (CENTRALIZADO)
# ============================================

# Importar funções de autenticação do módulo centralizado
from modules.auth import verificar_token, verificar_admin, verificar_token_cliente

# ============================================
# ROTAS MODULARES (Refatoração 23/01/2026)
# ============================================

# Importar e registrar routers dos módulos
from routes.auth import router as auth_router
from routes.prazos import router as prazos_router
from routes.datajud import router as datajud_router
from routes.banners import router as banners_router
from routes.admin_processos import router as admin_processos_router
from routes.portal_cliente import router as portal_cliente_router
from routes.admin_cadastros import router as admin_cadastros_router

# Registrar routers na aplicação
app.include_router(auth_router)
app.include_router(prazos_router)
app.include_router(datajud_router)
app.include_router(banners_router)
app.include_router(admin_processos_router)
app.include_router(portal_cliente_router)
app.include_router(admin_cadastros_router)

# ============================================
# INICIALIZAÇÃO DA APLICAÇÃO
# ============================================

@app.on_event("startup")
def startup():
    logger.info("Iniciando aplicação...")
    logger.info(f"RESEND_API_KEY configurada: {bool(RESEND_API_KEY)}")
    logger.info(f"FROM_EMAIL: {FROM_EMAIL}")
    init_db()

# ============================================
# FUNÇÕES DO BANCO - DEMANDAS ESPECÍFICAS
# ============================================

def salvar_dados_demanda(cadastro_id: str, tipo_demanda: str, dados: dict, status: str = "rascunho") -> bool:
    """Salva ou atualiza dados específicos de uma demanda."""
    conn = get_db()
    if not conn:
        return False
    
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO dados_demanda_especifica (cadastro_id, tipo_demanda, dados, status)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (cadastro_id, tipo_demanda) DO UPDATE SET
                dados = EXCLUDED.dados,
                status = EXCLUDED.status,
                atualizado_em = CURRENT_TIMESTAMP
        """, (cadastro_id, tipo_demanda, json.dumps(dados), status))
        conn.commit()
        cur.close()
        conn.close()
        return True
    except Exception as e:
        logger.error(f"Erro ao salvar dados da demanda: {e}")
        return False

def buscar_dados_demanda(cadastro_id: str, tipo_demanda: str) -> dict:
    """Busca dados específicos de uma demanda."""
    conn = get_db()
    if not conn:
        return None
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT * FROM dados_demanda_especifica 
            WHERE cadastro_id = %s AND tipo_demanda = %s
        """, (cadastro_id, tipo_demanda))
        row = cur.fetchone()
        cur.close()
        conn.close()
        
        if row:
            dados = row['dados']
            if isinstance(dados, str):
                dados = json.loads(dados)
            return {
                "id": row['id'],
                "cadastro_id": row['cadastro_id'],
                "tipo_demanda": row['tipo_demanda'],
                "dados": dados,
                "status": row['status'],
                "criado_em": row['criado_em'].isoformat() if row['criado_em'] else None,
                "atualizado_em": row['atualizado_em'].isoformat() if row['atualizado_em'] else None
            }
        return None
    except Exception as e:
        logger.error(f"Erro ao buscar dados da demanda: {e}")
        return None

def salvar_documento_demanda(cadastro_id: str, tipo_documento: str, nome_arquivo: str, 
                              nome_original: str, arquivo_path: str, descricao: str = "") -> bool:
    """Salva referência de documento específico da demanda."""
    conn = get_db()
    if not conn:
        return False
    
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO documentos_demanda (cadastro_id, tipo_documento, nome_arquivo, 
                nome_original, arquivo_path, descricao)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (cadastro_id, tipo_documento, nome_arquivo, nome_original, arquivo_path, descricao))
        conn.commit()
        cur.close()
        conn.close()
        return True
    except Exception as e:
        logger.error(f"Erro ao salvar documento da demanda: {e}")
        return False

def listar_documentos_demanda(cadastro_id: str) -> list:
    """Lista documentos específicos de uma demanda."""
    conn = get_db()
    if not conn:
        return []
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT * FROM documentos_demanda 
            WHERE cadastro_id = %s
            ORDER BY criado_em DESC
        """, (cadastro_id,))
        rows = cur.fetchall()
        cur.close()
        conn.close()
        
        return [{
            "id": row['id'],
            "tipo_documento": row['tipo_documento'],
            "nome_arquivo": row['nome_arquivo'],
            "nome_original": row['nome_original'],
            "descricao": row['descricao'],
            "arquivo_path": row.get('arquivo_path', ''),
            "criado_em": row['criado_em'].isoformat() if row['criado_em'] else None
        } for row in rows]
    except Exception as e:
        logger.error(f"Erro ao listar documentos da demanda: {e}")
        return []

def buscar_documento_demanda(doc_id: int) -> dict:
    """Busca um documento da demanda específico."""
    conn = get_db()
    if not conn:
        return None
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM documentos_demanda WHERE id = %s", (doc_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        return dict(row) if row else None
    except Exception as e:
        logger.error(f"Erro ao buscar documento da demanda: {e}")
        return None

# ============================================
# ROTAS DA API - BÁSICAS
# ============================================

@app.get("/")
def root():
    return {"message": "Vaucher e Álvares API", "status": "online", "version": "3.0"}

@app.api_route("/health", methods=["GET", "HEAD"])
def health():
    return {
        "status": "healthy",
        "database": "connected" if get_db() else "disconnected",
        "email": "resend" if RESEND_API_KEY else "not_configured"
    }

# --- CADASTROS ---

@app.post("/api/cadastros")
async def criar_cadastro(dados: DadosCliente):
    """Recebe novo cadastro do cliente."""
    logger.info(f"Novo cadastro recebido: {dados.nome}")

    novo_cadastro = {
        "id": uuid.uuid4().hex[:12],
        "data": datetime.now().isoformat(),
        "data_cadastro_br": datetime.now().strftime("%d/%m/%Y"),
        "status": "pendente",
        "dados": dados.dict(),
        "documentos": [],
        "arquivos_gerados": {}
    }
    
    if salvar_cadastro(novo_cadastro):
        logger.info(f"Cadastro salvo com ID: {novo_cadastro['id']}")
        
        try:
            conteudo = f"""
                <p style="font-size: 16px;">Prezado(a) <strong>{dados.nome}</strong>,</p>
                <p>Seu cadastro foi recebido com sucesso!</p>
                <p>Nossa equipe irá analisar as informações e documentos enviados. 
                Em breve você receberá o Contrato de Honorários e a Procuração 
                para assinatura.</p>
                
                <div style="background-color: #f0f0f0; padding: 20px; border-radius: 8px; margin: 20px 0;">
                    <p style="margin: 0 0 10px 0;"><strong>📋 Seu código de protocolo:</strong></p>
                    <p style="font-size: 20px; font-family: monospace; background: #fff; padding: 10px; border-radius: 5px; text-align: center; margin: 0; color: #8B1538; font-weight: bold;">{novo_cadastro['id']}</p>
                    <p style="margin: 10px 0 0 0; font-size: 12px; color: #666;">Guarde este código para acompanhar seu cadastro</p>
                </div>
                
                <div style="background-color: #e3f2fd; padding: 15px; border-radius: 5px; margin: 20px 0;">
                    <p style="margin: 0;"><strong>⏱ Prazo estimado:</strong> até 2 dias úteis</p>
                </div>
                
                <p>Agradecemos a confiança em nosso escritório!</p>
            """
            corpo_html = criar_email_html(conteudo)
            
            await enviar_email_resend(
                dados.email,
                "✅ Cadastro Recebido - Vaucher e Álvares Advogados",
                corpo_html
            )
        except Exception as e:
            logger.error(f"Erro ao enviar e-mail de confirmação: {e}")
        
        return {"success": True, "id": novo_cadastro["id"]}
    else:
        raise HTTPException(status_code=500, detail="Erro ao salvar cadastro")

@app.get("/api/cadastros")
def listar_cadastros():
    """Lista todos os cadastros (painel admin)."""
    return carregar_cadastros()

@app.get("/api/cadastros/{cadastro_id}")
def obter_cadastro(cadastro_id: str):
    """Obtém detalhes de um cadastro específico."""
    cadastro = buscar_cadastro(cadastro_id)
    if cadastro:
        return cadastro
    raise HTTPException(status_code=404, detail="Cadastro não encontrado")

@app.delete("/api/cadastros/{cadastro_id}")
def deletar_cadastro(cadastro_id: str, usuario: dict = Depends(verificar_admin)):
    """Deleta um cadastro permanentemente (apenas admin)."""
    logger.info(f"Deletando cadastro: {cadastro_id} por {usuario['email']}")
    
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão com banco")
    
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM cadastros WHERE id = %s", (cadastro_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        cliente_uploads = os.path.join(UPLOADS_DIR, cadastro_id)
        cliente_gerados = os.path.join(GERADOS_DIR, cadastro_id)
        
        if os.path.exists(cliente_uploads):
            shutil.rmtree(cliente_uploads)
        if os.path.exists(cliente_gerados):
            shutil.rmtree(cliente_gerados)
        
        logger.info(f"Cadastro {cadastro_id} deletado com sucesso")
        return {"success": True, "message": "Cadastro deletado com sucesso"}
    except Exception as e:
        logger.error(f"Erro ao deletar cadastro: {e}")
        raise HTTPException(status_code=500, detail="Erro ao deletar cadastro")

@app.put("/api/cadastros/{cadastro_id}/validar")
def validar_cadastro(cadastro_id: str):
    """Marca cadastro como validado."""
    if atualizar_status(cadastro_id, "validado"):
        return {"success": True}
    raise HTTPException(status_code=404, detail="Cadastro não encontrado")

@app.post("/api/cadastros/{cadastro_id}/gerar-documentos")
def gerar_documentos(cadastro_id: str):
    """Gera documentos SEM enviar por e-mail."""
    logger.info(f"Gerando documentos para cadastro: {cadastro_id}")
    
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    try:
        dados = cadastro["dados"]
        arquivos = gerador.gerar_todos(dados, cadastro_id)
        logger.info(f"Documentos gerados: {arquivos}")
        
        cadastro["status"] = "documentos_gerados"
        cadastro["arquivos_gerados"] = arquivos
        salvar_cadastro(cadastro)
        
        return {
            "success": True,
            "arquivos": arquivos,
            "message": "Documentos gerados com sucesso!"
        }
    except Exception as e:
        logger.error(f"Erro ao gerar documentos: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/cadastros/{cadastro_id}/enviar-email")
async def enviar_email_documentos(
    cadastro_id: str,
    assunto: str = Form(default="Seus Documentos - Vaucher e Álvares Advogados"),
    mensagem: str = Form(default=""),
    arquivos: List[UploadFile] = File(default=[])
):
    """Envia documentos por e-mail."""
    logger.info(f"Enviando e-mail para cadastro: {cadastro_id}")
    
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    dados = cadastro["dados"]
    anexos_email = []
    
    if arquivos:
        for arquivo in arquivos:
            if arquivo.filename:
                # Validar arquivo
                valido, erro = validar_arquivo(arquivo.filename, arquivo.size or 0)
                if not valido:
                    raise HTTPException(status_code=400, detail=f"Arquivo inválido: {erro}")

                logger.info(f"Processando anexo: {arquivo.filename}")
                conteudo = await arquivo.read()
                nome_seguro = sanitizar_nome_arquivo(arquivo.filename)
                anexos_email.append({
                    "filename": nome_seguro,
                    "content": base64.b64encode(conteudo).decode("utf-8")
                })
    
    if not anexos_email:
        raise HTTPException(status_code=400, detail="Nenhum arquivo selecionado para envio")
    
    PORTAL_URL = os.getenv("PORTAL_URL", "https://cadastro.vaucherealvares.com")
    link_envio = f"{PORTAL_URL}/enviar-assinados?id={cadastro_id}"
    
    mensagem_html = f"<p>{mensagem}</p>" if mensagem else ""
    
    conteudo = f"""
        <p style="font-size: 16px;">Prezado(a) <strong>{dados['nome']}</strong>,</p>
        <p>Seguem em anexo os documentos para sua análise e assinatura.</p>
        {mensagem_html}
        <p>Por favor, leia atentamente os documentos. Após assiná-los, devolva-os por uma das opções abaixo:</p>
        
        <div style="background-color: #f0f0f0; padding: 20px; border-radius: 8px; margin: 20px 0;">
            <p style="margin: 0 0 10px 0;"><strong>📋 Seu código de protocolo:</strong></p>
            <p style="font-size: 20px; font-family: monospace; background: #fff; padding: 10px; border-radius: 5px; text-align: center; margin: 0; color: #8B1538; font-weight: bold;">{cadastro_id}</p>
        </div>
        
        <div style="background-color: #e8f5e9; padding: 20px; border-radius: 8px; margin: 20px 0;">
            <p style="margin: 0 0 15px 0;"><strong>📤 Como devolver os documentos assinados:</strong></p>
            
            <p style="margin: 0 0 10px 0;"><strong>Opção 1 - Pelo nosso portal:</strong></p>
            <p style="margin: 0 0 15px 0;">
                <a href="{link_envio}" style="background-color: #8B1538; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; display: inline-block;">
                    Clique aqui para enviar os documentos assinados
                </a>
            </p>
            
            <p style="margin: 0 0 10px 0;"><strong>Opção 2 - Por e-mail:</strong></p>
            <p style="margin: 0;">Responda este e-mail com os documentos anexados ou envie para:<br>
            <a href="mailto:atendimento@vaucherealvares.com" style="color: #8B1538; font-weight: bold;">atendimento@vaucherealvares.com</a></p>
        </div>
        
        <div style="background-color: #fff3e0; padding: 15px; border-radius: 5px; margin: 20px 0;">
            <p style="margin: 0;"><strong>📎 Anexos neste e-mail:</strong> {len(anexos_email)} documento(s)</p>
        </div>
        
        <p><strong>Dúvidas?</strong> Entre em contato conosco pelos canais abaixo.</p>
    """
    corpo_html = criar_email_html(conteudo)
    
    sucesso = await enviar_email_resend(dados["email"], assunto, corpo_html, anexos_email)
    
    if sucesso:
        cadastro["status"] = "enviado"
        salvar_cadastro(cadastro)
        return {"success": True, "message": f"E-mail enviado para {dados['email']} com {len(anexos_email)} anexo(s)"}
    else:
        raise HTTPException(status_code=500, detail="Erro ao enviar e-mail. Verifique os logs.")


# ============================================
# FUNÇÕES DE EMAIL - ATUALIZAÇÃO CADASTRAL
# ============================================

def enviar_email_solicitacao_atualizacao(email: str, nome: str, motivo: str) -> bool:
    """Envia email ao cliente solicitando atualização cadastral."""
    import httpx
    
    try:
        html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background-color: #1e3a5f; padding: 20px; text-align: center;">
                <h1 style="color: white; margin: 0;">Vaucher e Álvares</h1>
                <p style="color: #ccc; margin: 5px 0 0 0;">Sociedade de Advogados</p>
            </div>
            <div style="padding: 30px; background-color: #f9f9f9;">
                <h2 style="color: #1e3a5f;">Solicitação de Atualização Cadastral</h2>
                <p>Prezado(a) <strong>{nome}</strong>,</p>
                <p>O escritório <strong>Vaucher e Álvares Sociedade de Advogados</strong> 
                solicita que você atualize seus dados cadastrais em nosso sistema.</p>
                {f'<div style="background-color: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; margin: 20px 0;"><strong>Motivo:</strong> {motivo}</div>' if motivo else ''}
                <p>Por favor, acesse o <strong>Portal do Cliente</strong> para enviar 
                os dados atualizados:</p>
                <p style="text-align: center; margin: 30px 0;">
                    <a href="https://appcliente.vaucherealvares.com" 
                       style="background-color: #1e3a5f; color: white; padding: 15px 40px; 
                              text-decoration: none; border-radius: 5px; display: inline-block;">
                        Acessar Portal do Cliente
                    </a>
                </p>
            </div>
            <div style="background-color: #eee; padding: 20px; text-align: center; font-size: 12px; color: #666;">
                <p style="margin: 0;">Vaucher e Álvares Sociedade de Advogados</p>
                <p style="margin: 5px 0;">Rua Lima, nº 106, Jardim das Américas - Cuiabá/MT</p>
                <p style="margin: 5px 0;">Tel: (65) 3025-1223</p>
            </div>
        </div>
        """
        
        with httpx.Client() as client:
            response = client.post(
                "https://api.resend.com/emails",
                headers={
                    "Authorization": f"Bearer {RESEND_API_KEY}",
                    "Content-Type": "application/json"
                },
                json={
                    "from": FROM_EMAIL,
                    "to": email,
                    "subject": "Solicitação de Atualização Cadastral - Vaucher e Álvares",
                    "html": html
                }
            )
        
        if response.status_code in [200, 201]:
            logger.info(f"Email de solicitação de atualização enviado para {email}")
            return True
        else:
            logger.error(f"Erro ao enviar email: {response.text}")
            return False
    except Exception as e:
        logger.error(f"Erro ao enviar email de solicitação: {e}")
        return False


def enviar_email_atualizacao_aprovada(email: str, nome: str) -> bool:
    """Notifica cliente que atualização foi aprovada."""
    import httpx
    
    try:
        html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background-color: #1e3a5f; padding: 20px; text-align: center;">
                <h1 style="color: white; margin: 0;">Vaucher e Álvares</h1>
            </div>
            <div style="padding: 30px; background-color: #f9f9f9;">
                <div style="background-color: #d4edda; border-left: 4px solid #28a745; padding: 15px; margin-bottom: 20px;">
                    <h2 style="color: #155724; margin: 0;">✓ Atualização Cadastral Aprovada</h2>
                </div>
                <p>Prezado(a) <strong>{nome}</strong>,</p>
                <p>Sua atualização cadastral foi <strong style="color: #28a745;">aprovada</strong> 
                e seus dados foram atualizados em nosso sistema.</p>
                <p>Agradecemos pela colaboração!</p>
            </div>
            <div style="background-color: #eee; padding: 20px; text-align: center; font-size: 12px; color: #666;">
                <p style="margin: 0;">Vaucher e Álvares Sociedade de Advogados</p>
            </div>
        </div>
        """
        
        with httpx.Client() as client:
            response = client.post(
                "https://api.resend.com/emails",
                headers={
                    "Authorization": f"Bearer {RESEND_API_KEY}",
                    "Content-Type": "application/json"
                },
                json={
                    "from": FROM_EMAIL,
                    "to": email,
                    "subject": "✓ Atualização Cadastral Aprovada - Vaucher e Álvares",
                    "html": html
                }
            )
        
        return response.status_code in [200, 201]
    except Exception as e:
        logger.error(f"Erro ao enviar email de aprovação: {e}")
        return False


def enviar_email_atualizacao_rejeitada(email: str, nome: str, motivo: str) -> bool:
    """Notifica cliente que atualização foi rejeitada."""
    import httpx
    
    try:
        html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background-color: #1e3a5f; padding: 20px; text-align: center;">
                <h1 style="color: white; margin: 0;">Vaucher e Álvares</h1>
            </div>
            <div style="padding: 30px; background-color: #f9f9f9;">
                <div style="background-color: #f8d7da; border-left: 4px solid #dc3545; padding: 15px; margin-bottom: 20px;">
                    <h2 style="color: #721c24; margin: 0;">Atualização Cadastral - Revisão Necessária</h2>
                </div>
                <p>Prezado(a) <strong>{nome}</strong>,</p>
                <p>Sua atualização cadastral precisa de ajustes antes de ser aprovada.</p>
                {f'<div style="background-color: #fff; border: 1px solid #ddd; padding: 15px; margin: 20px 0;"><strong>Observação:</strong> {motivo}</div>' if motivo else ''}
                <p>Por favor, acesse o Portal do Cliente e envie novamente os dados corrigidos.</p>
                <p style="text-align: center; margin: 30px 0;">
                    <a href="https://appcliente.vaucherealvares.com" 
                       style="background-color: #1e3a5f; color: white; padding: 15px 40px; 
                              text-decoration: none; border-radius: 5px; display: inline-block;">
                        Acessar Portal
                    </a>
                </p>
            </div>
            <div style="background-color: #eee; padding: 20px; text-align: center; font-size: 12px; color: #666;">
                <p style="margin: 0;">Vaucher e Álvares Sociedade de Advogados</p>
            </div>
        </div>
        """
        
        with httpx.Client() as client:
            response = client.post(
                "https://api.resend.com/emails",
                headers={
                    "Authorization": f"Bearer {RESEND_API_KEY}",
                    "Content-Type": "application/json"
                },
                json={
                    "from": FROM_EMAIL,
                    "to": email,
                    "subject": "Atualização Cadastral - Revisão Necessária - Vaucher e Álvares",
                    "html": html
                }
            )
        
        return response.status_code in [200, 201]
    except Exception as e:
        logger.error(f"Erro ao enviar email de rejeição: {e}")
        return False


def enviar_email_nova_atualizacao_admin(cadastro_id: str, nome_cliente: str) -> bool:
    """Notifica escritório sobre nova atualização cadastral."""
    import httpx
    
    try:
        html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background-color: #1e3a5f; padding: 20px; text-align: center;">
                <h1 style="color: white; margin: 0;">Nova Atualização Cadastral</h1>
            </div>
            <div style="padding: 30px; background-color: #f9f9f9;">
                <p>O cliente <strong>{nome_cliente}</strong> enviou uma atualização cadastral para análise.</p>
                <p><strong>ID do Cadastro:</strong> {cadastro_id}</p>
                <p style="text-align: center; margin: 30px 0;">
                    <a href="https://painel.vaucherealvares.com" 
                       style="background-color: #1e3a5f; color: white; padding: 15px 40px; 
                              text-decoration: none; border-radius: 5px; display: inline-block;">
                        Acessar Painel Administrativo
                    </a>
                </p>
            </div>
        </div>
        """
        
        with httpx.Client() as client:
            response = client.post(
                "https://api.resend.com/emails",
                headers={
                    "Authorization": f"Bearer {RESEND_API_KEY}",
                    "Content-Type": "application/json"
                },
                json={
                    "from": FROM_EMAIL,
                    "to": "atendimento@vaucherealvares.com",
                    "subject": f"Nova Atualização Cadastral - {nome_cliente}",
                    "html": html
                }
            )
        
        return response.status_code in [200, 201]
    except Exception as e:
        logger.error(f"Erro ao enviar email para admin: {e}")
        return False


@app.get("/api/cadastros/exportar/excel")
def exportar_cadastros_excel():
    """Exporta todos os cadastros para uma planilha Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="Biblioteca openpyxl não instalada")
    
    cadastros = carregar_cadastros()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cadastros"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8B1538", end_color="8B1538", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = [
        "ID", "Data Cadastro", "Status", "Nome", "CPF", "RG", 
        "Data Nascimento", "Estado Civil", "Nacionalidade", "Profissão",
        "Endereço", "E-mail", "Telefone", "Tipo de Demanda",
        "Objeto do Contrato", "Poderes Específicos", "Observações"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    tipos_demanda = {
        'adicional_insalubridade': 'Adicional de Insalubridade',
        'adicional_periculosidade': 'Adicional de Periculosidade',
        'desvio_funcao': 'Desvio de Função',
        'progressao_funcional': 'Progressão Funcional',
        'revisao_aposentadoria': 'Revisão de Aposentadoria',
        'licenca_premio': 'Licença Prêmio',
        'ferias_nao_gozadas': 'Férias Não Gozadas',
        'horas_extras': 'Horas Extras',
        'reintegracao': 'Reintegração',
        'outro': 'Outro'
    }
    
    status_map = {
        'pendente': 'Pendente',
        'validado': 'Validado',
        'documentos_gerados': 'Documentos Gerados',
        'enviado': 'Enviado',
        'assinado': 'Documentos Assinados Recebidos'
    }
    
    for row, cadastro in enumerate(cadastros, 2):
        dados = cadastro.get("dados", {})
        
        values = [
            cadastro.get("id", ""),
            cadastro.get("data", ""),
            status_map.get(cadastro.get("status", ""), cadastro.get("status", "")),
            dados.get("nome", ""),
            dados.get("cpf", ""),
            dados.get("rg", ""),
            dados.get("data_nascimento", ""),
            dados.get("estado_civil", ""),
            dados.get("nacionalidade", ""),
            dados.get("profissao", ""),
            dados.get("endereco_completo", ""),
            dados.get("email", ""),
            dados.get("telefone", ""),
            tipos_demanda.get(dados.get("tipo_demanda", ""), dados.get("tipo_demanda", "")),
            dados.get("objeto_contrato", ""),
            dados.get("poderes_especificos", ""),
            dados.get("observacoes", "")
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    column_widths = [12, 12, 15, 30, 15, 15, 12, 12, 12, 20, 40, 30, 15, 25, 50, 50, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.freeze_panes = "A2"
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"cadastros_vaucher_alvares_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/cadastros/{cadastro_id}/download/{tipo}")
def download_documento(cadastro_id: str, tipo: str):
    """Faz download de um documento gerado."""
    cadastro = buscar_cadastro(cadastro_id)
    
    if cadastro and cadastro.get("arquivos_gerados"):
        arquivo = cadastro["arquivos_gerados"].get(tipo)
        if arquivo and os.path.exists(arquivo):
            return FileResponse(
                arquivo,
                filename=os.path.basename(arquivo),
                media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            )
    
    raise HTTPException(status_code=404, detail="Documento não encontrado")

@app.post("/api/cadastros/{cadastro_id}/enviar-assinados")
async def receber_documentos_assinados(cadastro_id: str, arquivos: List[UploadFile] = File(...)):
    """Recebe documentos assinados enviados pelo cliente."""
    logger.info(f"ENVIAR-ASSINADOS: cadastro_id={cadastro_id}, arquivos={len(arquivos)}")
    
    # Buscar cadastro
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    # Criar pasta para documentos assinados (usando UPLOADS_DIR para consistência)
    pasta_assinados = os.path.join(UPLOADS_DIR, "documentos_assinados", cadastro_id)
    os.makedirs(pasta_assinados, exist_ok=True)

    # Salvar arquivos
    arquivos_salvos = []
    for arquivo in arquivos:
        # Sanitizar nome do arquivo para segurança
        nome_arquivo = sanitizar_nome_arquivo(arquivo.filename)
        caminho = os.path.join(pasta_assinados, nome_arquivo)
        
        with open(caminho, "wb") as f:
            conteudo = await arquivo.read()
            f.write(conteudo)
        
        arquivos_salvos.append(nome_arquivo)
        logger.info(f"Arquivo salvo: {caminho}")
    
    # Atualizar cadastro no banco
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão")
    
    try:
        cur = conn.cursor()
        
        # Pegar documentos assinados existentes
        docs_existentes = cadastro.get("documentos_assinados", []) or []
        docs_atualizados = docs_existentes + arquivos_salvos
        
        cur.execute("""
            UPDATE cadastros 
            SET documentos_assinados = %s,
                data_assinatura = NOW(),
                status = 'assinado'
            WHERE id = %s
        """, (json.dumps(docs_atualizados), cadastro_id))
        
        conn.commit()
        cur.close()
        conn.close()
        
        logger.info(f"Cadastro {cadastro_id} atualizado com {len(arquivos_salvos)} documentos assinados")
        
        return {
            "success": True,
            "message": f"{len(arquivos_salvos)} documento(s) recebido(s) com sucesso",
            "arquivos": arquivos_salvos
        }
    
    except Exception as e:
        logger.error(f"Erro ao salvar documentos assinados: {e}")
        raise HTTPException(status_code=500, detail="Erro ao salvar documentos")

# --- UPLOAD DE DOCUMENTOS DO CLIENTE ---

@app.post("/api/cadastros/{cadastro_id}/upload")
async def upload_documento(
    cadastro_id: str,
    arquivo: UploadFile = File(...),
    tipo_documento: str = Form(default=""),
    categoria: str = Form(default="")
):
    """Recebe upload de documento do cliente.

    Se tipo_documento ou categoria for especificado, salva na tabela documentos_demanda.
    Caso contrário, salva na lista de documentos genérica do cadastro.
    """
    # Validar arquivo
    valido, erro = validar_arquivo(arquivo.filename, arquivo.size or 0)
    if not valido:
        raise HTTPException(status_code=400, detail=f"Arquivo inválido: {erro}")

    logger.info(f"Upload recebido para cadastro {cadastro_id}: {arquivo.filename}, tipo: {tipo_documento}, categoria: {categoria}")

    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    # Determinar o tipo do documento
    tipo = tipo_documento or categoria
    
    if tipo:
        # Salvar como documento específico da demanda
        cliente_dir = os.path.join(UPLOADS_DIR, "documentos_demanda", cadastro_id)
        os.makedirs(cliente_dir, exist_ok=True)
        
        ext = os.path.splitext(arquivo.filename)[1]
        nome_arquivo = f"{tipo}_{uuid.uuid4().hex[:8]}{ext}"
        arquivo_path = os.path.join(cliente_dir, nome_arquivo)
        
        content = await arquivo.read()
        with open(arquivo_path, "wb") as f:
            f.write(content)
        
        # Salvar na tabela documentos_demanda
        sucesso = salvar_documento_demanda(
            cadastro_id, tipo, nome_arquivo, 
            arquivo.filename, arquivo_path, tipo
        )
        
        if not sucesso:
            logger.error(f"Erro ao salvar documento da demanda no banco: {arquivo.filename}")
        
        return {"success": True, "filename": arquivo.filename, "tipo": tipo, "modo": "demanda"}
    else:
        # Comportamento padrão - salva na pasta do cliente e lista genérica
        cliente_dir = os.path.join(UPLOADS_DIR, cadastro_id)
        os.makedirs(cliente_dir, exist_ok=True)

        nome_seguro = sanitizar_nome_arquivo(arquivo.filename)
        file_path = os.path.join(cliente_dir, nome_seguro)
        content = await arquivo.read()
        with open(file_path, "wb") as f:
            f.write(content)

        if nome_seguro not in cadastro["documentos"]:
            cadastro["documentos"].append(nome_seguro)
        salvar_cadastro(cadastro)

        return {"success": True, "filename": nome_seguro, "modo": "generico"}

@app.get("/api/cadastros/{cadastro_id}/uploads/{filename}")
def download_upload_cliente(cadastro_id: str, filename: str):
    """Faz download de um arquivo enviado pelo cliente."""
    # Sanitizar filename para prevenir path traversal
    nome_seguro = sanitizar_nome_arquivo(filename)
    file_path = os.path.join(UPLOADS_DIR, cadastro_id, nome_seguro)

    # Verificar se o caminho está dentro do diretório permitido
    real_path = os.path.realpath(file_path)
    allowed_dir = os.path.realpath(os.path.join(UPLOADS_DIR, cadastro_id))
    if not real_path.startswith(allowed_dir):
        raise HTTPException(status_code=403, detail="Acesso negado")

    if os.path.exists(file_path):
        return FileResponse(file_path, filename=nome_seguro, media_type="application/octet-stream")

    raise HTTPException(status_code=404, detail="Arquivo não encontrado")


@app.post("/api/cadastros/{cadastro_id}/uploads-categorizados")
async def upload_documentos_categorizados(
    cadastro_id: str,
    arquivos: List[UploadFile] = File(...),
    categorias: str = Form(default="")
):
    """Recebe múltiplos uploads com suas categorias.
    
    categorias deve ser uma string JSON com lista de categorias na mesma ordem dos arquivos.
    Exemplo: ["documentos_pessoais", "certificado_residencia", "contracheques"]
    
    Todos os documentos são salvos na tabela documentos_demanda.
    """
    logger.info(f"Uploads categorizados para cadastro {cadastro_id}: {len(arquivos)} arquivo(s)")
    
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    # Parse das categorias
    try:
        lista_categorias = json.loads(categorias) if categorias else []
    except:
        lista_categorias = []
    
    # Criar diretório
    cliente_dir = os.path.join(UPLOADS_DIR, "documentos_demanda", cadastro_id)
    os.makedirs(cliente_dir, exist_ok=True)
    
    salvos = []
    erros = []
    
    for i, arquivo in enumerate(arquivos):
        # Determinar categoria
        categoria = lista_categorias[i] if i < len(lista_categorias) else "documento_geral"
        
        try:
            ext = os.path.splitext(arquivo.filename)[1]
            nome_arquivo = f"{categoria}_{uuid.uuid4().hex[:8]}{ext}"
            arquivo_path = os.path.join(cliente_dir, nome_arquivo)
            
            content = await arquivo.read()
            with open(arquivo_path, "wb") as f:
                f.write(content)
            
            # Salvar na tabela documentos_demanda
            sucesso = salvar_documento_demanda(
                cadastro_id, categoria, nome_arquivo, 
                arquivo.filename, arquivo_path, categoria
            )
            
            if sucesso:
                salvos.append({"nome": arquivo.filename, "categoria": categoria})
            else:
                erros.append({"nome": arquivo.filename, "erro": "Erro ao salvar no banco"})
        except Exception as e:
            logger.error(f"Erro ao processar arquivo {arquivo.filename}: {e}")
            erros.append({"nome": arquivo.filename, "erro": str(e)})
    
    return {
        "success": len(salvos) > 0,
        "salvos": salvos,
        "erros": erros,
        "total_salvos": len(salvos),
        "total_erros": len(erros)
    }


# ============================================
# MÓDULO FINANCEIRO
# ============================================

@app.get("/api/cadastros/{cadastro_id}/financeiro")
def obter_financeiro(cadastro_id: str):
    """Obtém dados financeiros de um cadastro."""
    financeiro = buscar_financeiro(cadastro_id)
    if financeiro:
        return financeiro
    return {
        "cadastro_id": cadastro_id,
        "numero_processo": "",
        "vara_tribunal": "",
        "percentual_honorarios": 20,
        "valor_credito_cliente": 0,
        "depositos": [],
        "sucumbencias": [],
        "retencoes": [],
        "observacoes": ""
    }

@app.post("/api/cadastros/{cadastro_id}/financeiro")
def salvar_dados_financeiro(cadastro_id: str, dados: FinanceiroData):
    """Salva ou atualiza dados financeiros de um cadastro."""
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    if salvar_financeiro(cadastro_id, dados.dict()):
        return {"success": True, "message": "Dados financeiros salvos com sucesso"}
    raise HTTPException(status_code=500, detail="Erro ao salvar dados financeiros")

@app.get("/api/cadastros/{cadastro_id}/prestacao-contas")
def gerar_documento_prestacao_contas(cadastro_id: str):
    """Gera documento de prestação de contas."""
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")

    financeiro = buscar_financeiro(cadastro_id)
    if not financeiro:
        raise HTTPException(status_code=400, detail="Dados financeiros não encontrados.")

    depositos = financeiro.get("depositos", [])
    if not depositos or len(depositos) == 0:
        raise HTTPException(status_code=400, detail="Adicione pelo menos um depósito.")

    try:
        caminho_arquivo = gerador.gerar_prestacao_contas(
            dados_cliente=cadastro["dados"],
            financeiro=financeiro,
            cadastro_id=cadastro_id
        )

        # Atualizar status para 'gerado' e salvar caminho do arquivo
        atualizar_status_prestacao(cadastro_id, "gerado", caminho_arquivo)

        nome_arquivo = os.path.basename(caminho_arquivo)

        return FileResponse(
            caminho_arquivo,
            filename=nome_arquivo,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        )
    except Exception as e:
        logger.error(f"Erro ao gerar prestação de contas: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# ENDPOINTS CRUD - LANÇAMENTOS FINANCEIROS
# ============================================

class LancamentoItem(BaseModel):
    data: str = ""
    valor: float = 0
    descricao: str = ""
    origem: Optional[str] = ""


@app.post("/api/cadastros/{cadastro_id}/financeiro/depositos")
def adicionar_deposito(cadastro_id: str, item: LancamentoItem, usuario: dict = Depends(verificar_admin)):
    """Adiciona um depósito aos dados financeiros."""
    financeiro = buscar_financeiro(cadastro_id)
    if not financeiro:
        financeiro = {
            "numero_processo": "",
            "vara_tribunal": "",
            "percentual_honorarios": 20,
            "valor_credito_cliente": 0,
            "depositos": [],
            "sucumbencias": [],
            "retencoes": [],
            "observacoes": ""
        }

    depositos = financeiro.get("depositos", [])
    depositos.append({
        "data": item.data,
        "valor": item.valor,
        "descricao": item.descricao,
        "origem": item.origem or item.descricao
    })
    financeiro["depositos"] = depositos

    if salvar_financeiro(cadastro_id, financeiro):
        return {"success": True, "message": "Depósito adicionado", "depositos": depositos}
    raise HTTPException(status_code=500, detail="Erro ao salvar depósito")


@app.delete("/api/cadastros/{cadastro_id}/financeiro/depositos/{idx}")
def remover_deposito(cadastro_id: str, idx: int, usuario: dict = Depends(verificar_admin)):
    """Remove um depósito dos dados financeiros."""
    financeiro = buscar_financeiro(cadastro_id)
    if not financeiro:
        raise HTTPException(status_code=404, detail="Dados financeiros não encontrados")

    depositos = financeiro.get("depositos", [])
    if idx < 0 or idx >= len(depositos):
        raise HTTPException(status_code=400, detail="Índice inválido")

    depositos.pop(idx)
    financeiro["depositos"] = depositos

    if salvar_financeiro(cadastro_id, financeiro):
        return {"success": True, "message": "Depósito removido", "depositos": depositos}
    raise HTTPException(status_code=500, detail="Erro ao remover depósito")


@app.post("/api/cadastros/{cadastro_id}/financeiro/sucumbencias")
def adicionar_sucumbencia(cadastro_id: str, item: LancamentoItem, usuario: dict = Depends(verificar_admin)):
    """Adiciona uma sucumbência aos dados financeiros."""
    financeiro = buscar_financeiro(cadastro_id)
    if not financeiro:
        financeiro = {
            "numero_processo": "",
            "vara_tribunal": "",
            "percentual_honorarios": 20,
            "valor_credito_cliente": 0,
            "depositos": [],
            "sucumbencias": [],
            "retencoes": [],
            "observacoes": ""
        }

    sucumbencias = financeiro.get("sucumbencias", [])
    sucumbencias.append({
        "data": item.data,
        "valor": item.valor,
        "descricao": item.descricao
    })
    financeiro["sucumbencias"] = sucumbencias

    if salvar_financeiro(cadastro_id, financeiro):
        return {"success": True, "message": "Sucumbência adicionada", "sucumbencias": sucumbencias}
    raise HTTPException(status_code=500, detail="Erro ao salvar sucumbência")


@app.delete("/api/cadastros/{cadastro_id}/financeiro/sucumbencias/{idx}")
def remover_sucumbencia(cadastro_id: str, idx: int, usuario: dict = Depends(verificar_admin)):
    """Remove uma sucumbência dos dados financeiros."""
    financeiro = buscar_financeiro(cadastro_id)
    if not financeiro:
        raise HTTPException(status_code=404, detail="Dados financeiros não encontrados")

    sucumbencias = financeiro.get("sucumbencias", [])
    if idx < 0 or idx >= len(sucumbencias):
        raise HTTPException(status_code=400, detail="Índice inválido")

    sucumbencias.pop(idx)
    financeiro["sucumbencias"] = sucumbencias

    if salvar_financeiro(cadastro_id, financeiro):
        return {"success": True, "message": "Sucumbência removida", "sucumbencias": sucumbencias}
    raise HTTPException(status_code=500, detail="Erro ao remover sucumbência")


@app.post("/api/cadastros/{cadastro_id}/financeiro/retencoes")
def adicionar_retencao(cadastro_id: str, item: LancamentoItem, usuario: dict = Depends(verificar_admin)):
    """Adiciona uma retenção aos dados financeiros."""
    financeiro = buscar_financeiro(cadastro_id)
    if not financeiro:
        financeiro = {
            "numero_processo": "",
            "vara_tribunal": "",
            "percentual_honorarios": 20,
            "valor_credito_cliente": 0,
            "depositos": [],
            "sucumbencias": [],
            "retencoes": [],
            "observacoes": ""
        }

    retencoes = financeiro.get("retencoes", [])
    retencoes.append({
        "data": item.data,
        "valor": item.valor,
        "descricao": item.descricao
    })
    financeiro["retencoes"] = retencoes

    if salvar_financeiro(cadastro_id, financeiro):
        return {"success": True, "message": "Retenção adicionada", "retencoes": retencoes}
    raise HTTPException(status_code=500, detail="Erro ao salvar retenção")


@app.delete("/api/cadastros/{cadastro_id}/financeiro/retencoes/{idx}")
def remover_retencao(cadastro_id: str, idx: int, usuario: dict = Depends(verificar_admin)):
    """Remove uma retenção dos dados financeiros."""
    financeiro = buscar_financeiro(cadastro_id)
    if not financeiro:
        raise HTTPException(status_code=404, detail="Dados financeiros não encontrados")

    retencoes = financeiro.get("retencoes", [])
    if idx < 0 or idx >= len(retencoes):
        raise HTTPException(status_code=400, detail="Índice inválido")

    retencoes.pop(idx)
    financeiro["retencoes"] = retencoes

    if salvar_financeiro(cadastro_id, financeiro):
        return {"success": True, "message": "Retenção removida", "retencoes": retencoes}
    raise HTTPException(status_code=500, detail="Erro ao remover retenção")

# ÁREA DO CLIENTE - DEVOLUÇÃO DE DOCUMENTOS (EXISTENTE)
# ============================================

@app.get("/api/cliente/cadastro/{cadastro_id}")
def cliente_ver_cadastro_publico(cadastro_id: str):
    """Cliente visualiza seu próprio cadastro (sem autenticação, mas limitado)."""
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    return {
        "id": cadastro["id"],
        "nome": cadastro["dados"].get("nome", ""),
        "email": cadastro["dados"].get("email", ""),
        "status": cadastro["status"],
        "data": cadastro["data"],
        "tipo_demanda": cadastro["dados"].get("tipo_demanda", ""),
        "documentos_assinados": cadastro.get("documentos_assinados", [])
    }

@app.post("/api/cliente/{cadastro_id}/enviar-assinados")
async def cliente_enviar_documentos_assinados(
    cadastro_id: str,
    arquivos: List[UploadFile] = File(...)
):
    """Cliente envia documentos assinados."""
    logger.info(f"Cliente enviando documentos assinados para cadastro: {cadastro_id}")
    
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    if cadastro["status"] not in ["enviado", "assinado"]:
        raise HTTPException(status_code=400, detail="Você ainda não recebeu os documentos para assinar")
    
    # Usar caminho consistente com download: uploads/documentos_assinados/{cadastro_id}/
    cliente_assinados_dir = os.path.join(UPLOADS_DIR, "documentos_assinados", cadastro_id)
    os.makedirs(cliente_assinados_dir, exist_ok=True)
    
    arquivos_salvos = []
    
    for arquivo in arquivos:
        if arquivo.filename:
            nome_arquivo = f"ASSINADO_{arquivo.filename}"
            file_path = os.path.join(cliente_assinados_dir, nome_arquivo)
            
            with open(file_path, "wb") as f:
                content = await arquivo.read()
                f.write(content)
            
            arquivos_salvos.append(nome_arquivo)
            logger.info(f"Documento assinado salvo: {nome_arquivo}")
    
    if not arquivos_salvos:
        raise HTTPException(status_code=400, detail="Nenhum arquivo enviado")
    
    if "documentos_assinados" not in cadastro:
        cadastro["documentos_assinados"] = []
    
    cadastro["documentos_assinados"].extend(arquivos_salvos)
    cadastro["status"] = "assinado"
    cadastro["data_assinatura"] = datetime.now().isoformat()
    salvar_cadastro(cadastro)
    
    try:
        dados = cadastro["dados"]
        conteudo = f"""
            <p style="font-size: 16px;"><strong>📝 Novos documentos assinados recebidos!</strong></p>
            <p>O cliente <strong>{dados['nome']}</strong> enviou os documentos assinados.</p>
            <div style="background-color: #f0f0f0; padding: 15px; border-radius: 5px; margin: 20px 0;">
                <p style="margin: 0;"><strong>Cadastro:</strong> {cadastro_id}</p>
                <p style="margin: 0;"><strong>E-mail:</strong> {dados['email']}</p>
                <p style="margin: 0;"><strong>Telefone:</strong> {dados['telefone']}</p>
                <p style="margin: 0;"><strong>Arquivos:</strong> {len(arquivos_salvos)} documento(s)</p>
            </div>
            <p>Acesse o painel administrativo para visualizar os documentos.</p>
        """
        corpo_html = criar_email_html(conteudo)
        
        await enviar_email_resend(
            FROM_EMAIL,
            f"📝 Documentos Assinados - {dados['nome']}",
            corpo_html
        )
    except Exception as e:
        logger.error(f"Erro ao notificar escritório: {e}")
    
    return {
        "success": True,
        "message": f"{len(arquivos_salvos)} documento(s) enviado(s) com sucesso!",
        "arquivos": arquivos_salvos
    }

@app.get("/api/cadastros/{cadastro_id}/assinados/{filename}")
def download_documento_assinado(cadastro_id: str, filename: str):
    """Faz download de um documento assinado pelo cliente."""
    # Sanitizar filename para segurança
    nome_seguro = sanitizar_nome_arquivo(filename)

    # Caminho principal: uploads/documentos_assinados/{cadastro_id}/{filename}
    file_path = os.path.join(UPLOADS_DIR, "documentos_assinados", cadastro_id, nome_seguro)

    if os.path.exists(file_path):
        return FileResponse(file_path, filename=nome_seguro, media_type="application/octet-stream")

    # Fallback: tentar caminhos alternativos para arquivos legados
    caminhos_alternativos = [
        # Caminho antigo: uploads/{cadastro_id}/assinados/{filename}
        os.path.join(UPLOADS_DIR, cadastro_id, "assinados", nome_seguro),
        os.path.join(UPLOADS_DIR, cadastro_id, "assinados", filename),
        # Tentar com nome original sem sanitizar
        os.path.join(UPLOADS_DIR, "documentos_assinados", cadastro_id, filename),
        # Caminho /app/ hardcoded antigo
        f"/app/uploads/documentos_assinados/{cadastro_id}/{nome_seguro}",
        f"/app/uploads/{cadastro_id}/assinados/{nome_seguro}",
    ]

    for caminho in caminhos_alternativos:
        if os.path.exists(caminho):
            logger.info(f"Documento assinado encontrado em caminho alternativo: {caminho}")
            return FileResponse(caminho, filename=nome_seguro or filename, media_type="application/octet-stream")

    logger.error(f"Documento assinado não encontrado: {file_path}, tentativas: {caminhos_alternativos}")
    raise HTTPException(status_code=404, detail="Arquivo não encontrado")


# ============================================
# FUNÇÕES DO BANCO - PORTAL DO CLIENTE
# ============================================
# Movidas para modules/database.py em 24/01/2026:
# - criar_cliente_auth, buscar_cliente_auth, buscar_cliente_por_email
# - atualizar_senha_cliente, registrar_acesso_cliente
# - criar_processo, listar_processos, buscar_processo, atualizar_processo, deletar_processo
# - criar_andamento_processo, listar_andamentos_processo, deletar_andamento_processo
# - criar_contrato_honorarios, listar_contratos, buscar_contrato, atualizar_contrato, deletar_contrato
# - listar_parcelas, atualizar_parcela, marcar_parcela_paga
# - criar_comprovante, listar_comprovantes_pendentes, aprovar_comprovante, rejeitar_comprovante
# - criar_documento_admin, listar_documentos_admin, buscar_documento_admin, deletar_documento_admin
# - criar_documento_extra, listar_documentos_extras, buscar_documento_extra, deletar_documento_extra
# - buscar_processo_info, salvar_processo_info
# - listar_andamentos, criar_andamento, deletar_andamento
# - listar_mensagens, criar_mensagem, marcar_mensagens_lidas, contar_mensagens_nao_lidas



# ============================================
# VERIFICAÇÃO DE TOKEN DO CLIENTE
# ============================================

def verificar_token_cliente(authorization: str = Header(None)) -> dict:
    """Verifica se o token de cliente é válido."""
    if not authorization:
        raise HTTPException(status_code=401, detail="Token não fornecido")
    
    token = authorization.replace("Bearer ", "")
    
    cliente = decodificar_token_cliente(token)
    if not cliente:
        raise HTTPException(status_code=401, detail="Token inválido ou expirado")
    
    auth = buscar_cliente_auth(cliente["cadastro_id"])
    if not auth or not auth.get("ativo", True):
        raise HTTPException(status_code=401, detail="Cliente não encontrado ou inativo")
    
    return {
        "cadastro_id": cliente["cadastro_id"],
        "email": auth["email"],
        "nome": auth["nome"],
        "primeiro_acesso": auth.get("primeiro_acesso", False)
    }


# ============================================
# ENDPOINTS - AUTENTICAÇÃO DO CLIENTE
# ============================================
# Movidos para routes/portal_cliente.py em 24/01/2026:
# - /api/cliente/login
# - /api/cliente/alterar-senha


# ============================================
# ENDPOINTS - PORTAL DO CLIENTE
# ============================================
# Movidos para routes/portal_cliente.py em 24/01/2026:
# - /api/cliente/meus-dados
# - /api/cliente/meus-processos
# - /api/cliente/processo/{processo_id}/andamentos
# - /api/cliente/meus-contratos
# - /api/cliente/andamentos
# - /api/cliente/mensagens (GET, POST)
# - /api/cliente/mensagens/nao-lidas
# - /api/cliente/documentos-extras (POST)
# - /api/cliente/meus-documentos-extras
# - /api/cliente/documentos-extras/{doc_id} (DELETE)
# - /api/cliente/parcelas/{parcela_id}/comprovante


@app.get("/api/cliente/documentos")
async def portal_cliente_documentos(cliente: dict = Depends(verificar_token_cliente)):
    """Lista documentos disponíveis para o cliente (recebidos, enviados e da demanda)."""
    cadastro = buscar_cadastro(cliente["cadastro_id"])
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    documentos = []
    
    # 1. Documentos RECEBIDOS do Escritório (admin -> cliente)
    docs_admin = listar_documentos_admin(cliente["cadastro_id"])
    for doc in docs_admin:
        documentos.append({
            "id": doc['id'],
            "tipo": f"admin_{doc['id']}",
            "nome": doc["nome_original"],
            "descricao": doc.get("descricao", ""),
            "categoria": "Recebido do Escritório",
            "origem": "recebido",
            "disponivel": True,
            "data": doc["criado_em"].isoformat() if doc.get("criado_em") else None
        })
    
    # 2. Documentos ENVIADOS pelo Cliente (documentos extras)
    docs_extras = listar_documentos_extras(cliente["cadastro_id"])
    for doc in docs_extras:
        documentos.append({
            "id": doc['id'],
            "tipo": f"extra_{doc['id']}",
            "nome": doc["nome_original"],
            "descricao": doc.get("descricao", ""),
            "categoria": "Enviado por Você",
            "origem": "enviado",
            "disponivel": True,
            "data": doc["criado_em"].isoformat() if doc.get("criado_em") else None
        })
    
    # 3. Documentos da DEMANDA (enviados pelo cliente no cadastro da demanda)
    docs_demanda = listar_documentos_demanda(cliente["cadastro_id"])
    for doc in docs_demanda:
        # Traduzir tipo de documento para nome amigável
        tipos_nomes = {
            "processo_anterior": "Processo Anterior",
            "historico_financeiro": "Histórico Financeiro",
            "certificado_residencia": "Certificado de Residência",
            "doc_pessoais": "Documento Pessoal",
            "comprovante_residencia": "Comprovante de Residência",
            "contracheque": "Contracheque",
            "outros": "Outros"
        }
        tipo_nome = tipos_nomes.get(doc["tipo_documento"], doc["tipo_documento"])

        documentos.append({
            "id": doc['id'],
            "tipo": f"demanda_{doc['id']}",
            "nome": doc["nome_original"],
            "descricao": doc.get("descricao") or tipo_nome,
            "categoria": "Documento da Demanda",
            "origem": "enviado",
            "tipo_documento": doc["tipo_documento"],
            "disponivel": True,
            "data": doc.get("criado_em")
        })

    # 4. Documentos do CADASTRO (armazenados no campo JSONB dados.documentos)
    dados = cadastro.get("dados", {})
    if isinstance(dados, str):
        import json
        try:
            dados = json.loads(dados)
        except:
            dados = {}

    docs_cadastro = dados.get("documentos", [])
    if docs_cadastro:
        # Traduzir tipos de documentos do cadastro
        tipos_cadastro = {
            "rg_frente": "RG (Frente)",
            "rg_verso": "RG (Verso)",
            "cpf": "CPF",
            "comprovante_residencia": "Comprovante de Residência",
            "contracheque": "Contracheque",
            "certidao_casamento": "Certidão de Casamento",
            "certidao_nascimento": "Certidão de Nascimento",
            "documento_militar": "Documento Militar",
            "procuracao": "Procuração",
            "outros": "Outros Documentos"
        }

        for idx, doc in enumerate(docs_cadastro):
            tipo_doc = doc.get("tipo", "documento")
            nome_amigavel = tipos_cadastro.get(tipo_doc, tipo_doc.replace("_", " ").title())

            documentos.append({
                "id": f"cadastro_{idx}",
                "tipo": f"cadastro_{tipo_doc}_{idx}",
                "nome": doc.get("nome") or nome_amigavel,
                "descricao": nome_amigavel,
                "categoria": "Documento do Cadastro",
                "origem": "enviado",
                "url": doc.get("url", ""),
                "disponivel": bool(doc.get("url")),
                "data": cadastro.get("data")
            })

    # Ordenar por data (mais recente primeiro)
    documentos.sort(key=lambda x: x.get("data") or "", reverse=True)

    return {"documentos": documentos}


@app.get("/api/cliente/documentos/{tipo}/download")
async def portal_cliente_download_documento(
    tipo: str,
    cliente: dict = Depends(verificar_token_cliente)
):
    """Cliente baixa um documento."""
    cadastro = buscar_cadastro(cliente["cadastro_id"])
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    # Documento do admin (recebido do escritório)
    if tipo.startswith("admin_"):
        doc_id = int(tipo.replace("admin_", ""))
        doc = buscar_documento_admin(doc_id)
        if not doc or doc["cadastro_id"] != cliente["cadastro_id"]:
            raise HTTPException(status_code=404, detail="Documento não encontrado")
        
        if not os.path.exists(doc["arquivo_path"]):
            raise HTTPException(status_code=404, detail="Arquivo não encontrado")
        
        return FileResponse(
            doc["arquivo_path"],
            filename=doc["nome_original"],
            media_type="application/octet-stream"
        )
    
    # Documento extra (enviado pelo cliente)
    if tipo.startswith("extra_"):
        doc_id = int(tipo.replace("extra_", ""))
        doc = buscar_documento_extra(doc_id)
        if not doc or doc["cadastro_id"] != cliente["cadastro_id"]:
            raise HTTPException(status_code=404, detail="Documento não encontrado")
        
        if not os.path.exists(doc["arquivo_path"]):
            raise HTTPException(status_code=404, detail="Arquivo não encontrado")
        
        return FileResponse(
            doc["arquivo_path"],
            filename=doc["nome_original"],
            media_type="application/octet-stream"
        )
    
    # Documento da demanda (enviado pelo cliente no cadastro)
    if tipo.startswith("demanda_"):
        doc_id = int(tipo.replace("demanda_", ""))
        doc = buscar_documento_demanda(doc_id)
        if not doc or doc["cadastro_id"] != cliente["cadastro_id"]:
            raise HTTPException(status_code=404, detail="Documento não encontrado")
        
        if not doc.get("arquivo_path") or not os.path.exists(doc["arquivo_path"]):
            raise HTTPException(status_code=404, detail="Arquivo não encontrado")
        
        return FileResponse(
            doc["arquivo_path"],
            filename=doc["nome_original"],
            media_type="application/octet-stream"
        )
    
    # Documento do CADASTRO (armazenado no campo JSONB dados.documentos)
    if tipo.startswith("cadastro_"):
        dados = cadastro.get("dados", {})
        if isinstance(dados, str):
            import json
            try:
                dados = json.loads(dados)
            except:
                dados = {}

        docs_cadastro = dados.get("documentos", [])

        # Extrair índice do tipo (cadastro_tipo_idx)
        partes = tipo.split("_")
        if len(partes) >= 3:
            try:
                idx = int(partes[-1])
                if 0 <= idx < len(docs_cadastro):
                    doc = docs_cadastro[idx]
                    url = doc.get("url", "")

                    if url:
                        # Se a URL for um arquivo local
                        if url.startswith("/") or url.startswith("./"):
                            # Construir caminho de forma segura
                            if url.startswith("/app/uploads/"):
                                # Extrair caminho relativo após /app/uploads/
                                relative_path = url[len("/app/uploads/"):]
                                arquivo_path = os.path.join(UPLOADS_DIR, relative_path)
                            elif url.startswith("/app/"):
                                # Outros caminhos /app/
                                relative_path = url[len("/app/"):]
                                arquivo_path = os.path.join(BASE_DIR, relative_path)
                            elif url.startswith("./"):
                                arquivo_path = os.path.join(BASE_DIR, url[2:])
                            else:
                                arquivo_path = url

                            if os.path.exists(arquivo_path):
                                nome = doc.get("nome") or os.path.basename(arquivo_path)
                                return FileResponse(
                                    arquivo_path,
                                    filename=nome,
                                    media_type="application/octet-stream"
                                )
                        # Se for URL externa, fazer proxy do download
                        elif url.startswith("http"):
                            import httpx
                            try:
                                async with httpx.AsyncClient() as client:
                                    resp = await client.get(url, follow_redirects=True)
                                    if resp.status_code == 200:
                                        nome = doc.get("nome") or url.split("/")[-1]
                                        return StreamingResponse(
                                            iter([resp.content]),
                                            media_type="application/octet-stream",
                                            headers={"Content-Disposition": f'attachment; filename="{nome}"'}
                                        )
                            except Exception as e:
                                logger.error(f"Erro ao baixar documento externo: {e}")
                                raise HTTPException(status_code=500, detail="Erro ao baixar documento externo")

            except (ValueError, IndexError):
                pass

        raise HTTPException(status_code=404, detail="Documento do cadastro não encontrado")

    # Documento gerado (contrato/procuração) - mantido para compatibilidade
    arquivos_gerados = cadastro.get("arquivos_gerados", {})
    if tipo in ["contrato", "procuracao"]:
        arquivo_path = arquivos_gerados.get(tipo)
        if not arquivo_path or not os.path.exists(arquivo_path):
            raise HTTPException(status_code=404, detail="Documento não encontrado")

        nome = "Contrato de Honorários.docx" if tipo == "contrato" else "Procuração.docx"
        return FileResponse(
            arquivo_path,
            filename=nome,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        )

    raise HTTPException(status_code=404, detail="Tipo de documento não reconhecido")


@app.post("/api/admin/clientes/lote")
async def admin_criar_clientes_lote(
    arquivo: UploadFile = File(...),
    habilitar_portal: bool = Form(default=False),
    usuario: dict = Depends(verificar_admin)
):
    """
    Admin cria múltiplos clientes a partir de arquivo TXT.

    Formato esperado do arquivo (separado por ponto e vírgula ou tabulação):
    nome;email;cpf;telefone;endereco;profissao;estado_civil;data_nascimento

    Exemplo:
    João Silva;joao@email.com;123.456.789-00;(65)99999-9999;Rua A, 123;Advogado;Casado;01/01/1990
    Maria Santos;maria@email.com;987.654.321-00;(65)88888-8888;Rua B, 456;Médica;Solteira;15/05/1985
    """
    logger.info(f"Admin {usuario['email']} importando clientes em lote")

    if not arquivo.filename.endswith(('.txt', '.csv')):
        raise HTTPException(status_code=400, detail="Formato de arquivo inválido. Use .txt ou .csv")

    # Ler conteúdo do arquivo
    try:
        conteudo = await arquivo.read()
        # Tentar decodificar com diferentes encodings
        for encoding in ['utf-8', 'latin-1', 'cp1252']:
            try:
                texto = conteudo.decode(encoding)
                break
            except:
                continue
        else:
            raise HTTPException(status_code=400, detail="Não foi possível ler o arquivo. Verifique a codificação.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler arquivo: {str(e)}")

    # Processar linhas
    linhas = texto.strip().split('\n')
    if len(linhas) < 1:
        raise HTTPException(status_code=400, detail="Arquivo vazio")

    # Detectar separador (ponto e vírgula ou tabulação)
    primeira_linha = linhas[0]
    separador = ';' if ';' in primeira_linha else '\t' if '\t' in primeira_linha else ','

    # Verificar se primeira linha é cabeçalho
    cabecalho_possivel = primeira_linha.lower()
    tem_cabecalho = 'nome' in cabecalho_possivel or 'email' in cabecalho_possivel or 'cpf' in cabecalho_possivel

    if tem_cabecalho:
        linhas = linhas[1:]  # Pular cabeçalho

    # Carregar cadastros existentes para verificar duplicatas
    cadastros_existentes = carregar_cadastros()
    cpfs_existentes = set()
    for c in cadastros_existentes:
        cpf = c.get("dados", {}).get("cpf", "").replace(".", "").replace("-", "")
        if cpf:
            cpfs_existentes.add(cpf)

    resultados = {
        "total": len(linhas),
        "criados": 0,
        "erros": 0,
        "detalhes": [],
        "clientes_criados": []
    }

    for i, linha in enumerate(linhas, 1):
        linha = linha.strip()
        if not linha:
            continue

        campos = linha.split(separador)

        # Mínimo: nome, email, cpf, telefone
        if len(campos) < 4:
            resultados["erros"] += 1
            resultados["detalhes"].append({
                "linha": i,
                "erro": f"Linha incompleta. Esperado no mínimo: nome;email;cpf;telefone"
            })
            continue

        try:
            nome = campos[0].strip()
            email = campos[1].strip()
            cpf = campos[2].strip()
            telefone = campos[3].strip()
            endereco = campos[4].strip() if len(campos) > 4 else ""
            profissao = campos[5].strip() if len(campos) > 5 else ""
            estado_civil = campos[6].strip() if len(campos) > 6 else ""
            data_nascimento = campos[7].strip() if len(campos) > 7 else ""

            # Validar campos obrigatórios
            if not nome or not email or not cpf or not telefone:
                resultados["erros"] += 1
                resultados["detalhes"].append({
                    "linha": i,
                    "erro": "Campos obrigatórios vazios (nome, email, cpf, telefone)"
                })
                continue

            # Verificar duplicata
            cpf_limpo = cpf.replace(".", "").replace("-", "")
            if cpf_limpo in cpfs_existentes:
                resultados["erros"] += 1
                resultados["detalhes"].append({
                    "linha": i,
                    "nome": nome,
                    "erro": f"CPF {cpf} já cadastrado"
                })
                continue

            # Criar cadastro
            novo_cadastro = {
                "id": uuid.uuid4().hex[:12],
                "data": datetime.now().isoformat(),
                "data_cadastro_br": datetime.now().strftime("%d/%m/%Y"),
                "status": "validado",
                "dados": {
                    "nome": nome,
                    "email": email,
                    "cpf": cpf,
                    "telefone": telefone,
                    "endereco_completo": endereco,
                    "profissao": profissao,
                    "estado_civil": estado_civil,
                    "data_nascimento": data_nascimento,
                    "nacionalidade": "brasileiro(a)",
                    "tipo_demanda": "",
                    "objeto_contrato": "",
                    "poderes_especificos": "",
                },
                "documentos": [],
                "arquivos_gerados": {}
            }

            if salvar_cadastro(novo_cadastro):
                cpfs_existentes.add(cpf_limpo)
                resultados["criados"] += 1

                cliente_info = {
                    "id": novo_cadastro["id"],
                    "nome": nome,
                    "email": email
                }

                # Habilitar portal se solicitado
                if habilitar_portal:
                    senha_temporaria = secrets.token_urlsafe(8)
                    if criar_cliente_auth(novo_cadastro["id"], senha_temporaria):
                        cliente_info["portal_habilitado"] = True
                        cliente_info["senha_temporaria"] = senha_temporaria

                        # Enviar email com credenciais
                        if RESEND_API_KEY:
                            try:
                                conteudo_email = f"""
                                    <p style="font-size: 16px;">Olá, <strong>{nome}</strong>!</p>

                                    <p>Seu cadastro foi criado e seu acesso ao Portal do Cliente foi habilitado.</p>

                                    <div style="background-color: #f5f5f5; padding: 20px; border-radius: 8px; margin: 20px 0;">
                                        <p style="margin: 0;"><strong>Email:</strong> {email}</p>
                                        <p style="margin: 10px 0;"><strong>Senha temporária:</strong> <code style="background: #e0e0e0; padding: 3px 8px; border-radius: 4px; font-size: 18px;">{senha_temporaria}</code></p>
                                    </div>

                                    <p>Acesse o portal em: <a href="https://appcliente.vaucherealvares.com" style="color: #8B1538;">appcliente.vaucherealvares.com</a></p>
                                """
                                corpo_html = criar_email_html(conteudo_email)
                                await enviar_email_resend(
                                    email,
                                    "🔑 Acesso ao Portal do Cliente - Vaucher e Álvares",
                                    corpo_html
                                )
                                cliente_info["email_enviado"] = True
                            except Exception as e:
                                logger.error(f"Erro ao enviar email para {email}: {e}")
                                cliente_info["email_enviado"] = False

                resultados["clientes_criados"].append(cliente_info)
                resultados["detalhes"].append({
                    "linha": i,
                    "nome": nome,
                    "sucesso": True,
                    "id": novo_cadastro["id"]
                })
            else:
                resultados["erros"] += 1
                resultados["detalhes"].append({
                    "linha": i,
                    "nome": nome,
                    "erro": "Erro ao salvar no banco de dados"
                })

        except Exception as e:
            resultados["erros"] += 1
            resultados["detalhes"].append({
                "linha": i,
                "erro": str(e)
            })

    logger.info(f"Importação em lote: {resultados['criados']} criados, {resultados['erros']} erros")

    return {
        "success": True,
        "message": f"{resultados['criados']} clientes criados, {resultados['erros']} erros",
        **resultados
    }



@app.get("/api/admin/comprovantes/{comprovante_id}/download")
async def admin_download_comprovante(
    comprovante_id: int,
    usuario: dict = Depends(verificar_admin)
):
    """Admin baixa um comprovante."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT arquivo_nome, arquivo_path FROM comprovantes WHERE id = %s", (comprovante_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        
        if not row:
            raise HTTPException(status_code=404, detail="Comprovante não encontrado")
        
        if not os.path.exists(row["arquivo_path"]):
            raise HTTPException(status_code=404, detail="Arquivo não encontrado")
        
        return FileResponse(
            row["arquivo_path"],
            filename=row["arquivo_nome"],
            media_type="application/octet-stream"
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# ENDPOINTS - ATUALIZAÇÃO CADASTRAL (ADMIN)
# ============================================

@app.post("/api/admin/clientes/{cadastro_id}/solicitar-atualizacao")
async def solicitar_atualizacao_cadastral(
    cadastro_id: str,
    dados: SolicitacaoAtualizacao,
    admin = Depends(verificar_admin)
):
    """Admin solicita que cliente atualize seus dados cadastrais."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Verificar se cadastro existe
        cur.execute("SELECT id FROM cadastros WHERE id = %s", (cadastro_id,))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Cadastro não encontrado")
        
        # Verificar se já existe solicitação pendente ou enviada
        cur.execute("""
            SELECT id FROM atualizacoes_cadastrais 
            WHERE cadastro_id = %s AND status IN ('pendente', 'enviada')
        """, (cadastro_id,))
        
        if cur.fetchone():
            raise HTTPException(
                status_code=400, 
                detail="Já existe uma solicitação de atualização pendente para este cliente"
            )
        
        # Criar solicitação
        cur.execute("""
            INSERT INTO atualizacoes_cadastrais 
            (cadastro_id, tipo, status, motivo_solicitacao, solicitado_em, solicitado_por)
            VALUES (%s, 'solicitada', 'pendente', %s, CURRENT_TIMESTAMP, %s)
            RETURNING id
        """, (cadastro_id, dados.motivo, admin["email"]))
        
        atualizacao_id = cur.fetchone()["id"]
        conn.commit()
        
        # Buscar dados do cliente para enviar email
        cur.execute("""
            SELECT dados->>'nome' as nome, dados->>'email' as email 
            FROM cadastros WHERE id = %s
        """, (cadastro_id,))
        cliente = cur.fetchone()
        
        email_enviado = False
        if cliente and cliente["email"]:
            email_enviado = enviar_email_solicitacao_atualizacao(
                cliente["email"], 
                cliente["nome"] or "Cliente", 
                dados.motivo
            )
        
        cur.close()
        conn.close()
        
        return {
            "success": True, 
            "id": atualizacao_id, 
            "message": "Solicitação de atualização criada com sucesso",
            "email_enviado": email_enviado
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao solicitar atualização: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/atualizacoes-pendentes")
async def listar_atualizacoes_pendentes(admin = Depends(verificar_admin)):
    """Lista todas as atualizações cadastrais pendentes de análise do admin."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT 
                a.id, a.cadastro_id, a.tipo, a.status,
                a.motivo_solicitacao, a.solicitado_em,
                a.dados_novos, a.documentos_novos, a.enviado_em,
                c.dados->>'nome' as nome_cliente,
                c.dados->>'email' as email_cliente,
                c.dados->>'cpf' as cpf_cliente
            FROM atualizacoes_cadastrais a
            JOIN cadastros c ON c.id = a.cadastro_id
            WHERE a.status = 'enviada'
            ORDER BY a.enviado_em DESC
        """)
        
        atualizacoes = []
        for row in cur.fetchall():
            atualizacao = dict(row)
            for campo in ['solicitado_em', 'enviado_em']:
                if atualizacao.get(campo):
                    atualizacao[campo] = atualizacao[campo].isoformat()
            atualizacoes.append(atualizacao)
        
        cur.close()
        conn.close()
        
        return {"atualizacoes": atualizacoes, "total": len(atualizacoes)}
        
    except Exception as e:
        logger.error(f"Erro ao listar atualizações: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/clientes/{cadastro_id}/atualizacoes")
async def listar_atualizacoes_cliente(cadastro_id: str, admin = Depends(verificar_admin)):
    """Lista histórico de atualizações cadastrais de um cliente específico."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, tipo, status, motivo_solicitacao, solicitado_em, 
                   solicitado_por, dados_novos, enviado_em, analisado_em, 
                   analisado_por, motivo_rejeicao, criado_em
            FROM atualizacoes_cadastrais 
            WHERE cadastro_id = %s 
            ORDER BY criado_em DESC
        """, (cadastro_id,))
        
        atualizacoes = []
        for row in cur.fetchall():
            atualizacao = dict(row)
            for campo in ['solicitado_em', 'enviado_em', 'analisado_em', 'criado_em']:
                if atualizacao.get(campo):
                    atualizacao[campo] = atualizacao[campo].isoformat()
            atualizacoes.append(atualizacao)
        
        cur.close()
        conn.close()
        
        return {"atualizacoes": atualizacoes}
        
    except Exception as e:
        logger.error(f"Erro ao listar atualizações do cliente: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/atualizacoes/{atualizacao_id}")
async def buscar_atualizacao(atualizacao_id: int, admin = Depends(verificar_admin)):
    """Busca detalhes de uma atualização específica."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT 
                a.*,
                c.dados as dados_atuais,
                c.dados->>'nome' as nome_cliente
            FROM atualizacoes_cadastrais a
            JOIN cadastros c ON c.id = a.cadastro_id
            WHERE a.id = %s
        """, (atualizacao_id,))
        
        result = cur.fetchone()
        if not result:
            raise HTTPException(status_code=404, detail="Atualização não encontrada")
        
        atualizacao = dict(result)
        for campo in ['solicitado_em', 'enviado_em', 'analisado_em', 'criado_em', 'atualizado_em']:
            if atualizacao.get(campo):
                atualizacao[campo] = atualizacao[campo].isoformat()
        
        cur.close()
        conn.close()
        
        return atualizacao
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao buscar atualização: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/admin/atualizacoes/{atualizacao_id}/aprovar")
async def aprovar_atualizacao(atualizacao_id: int, admin = Depends(verificar_admin)):
    """Aprova a atualização e aplica os novos dados ao cadastro do cliente."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT cadastro_id, dados_novos FROM atualizacoes_cadastrais 
            WHERE id = %s AND status = 'enviada'
        """, (atualizacao_id,))
        
        atualizacao = cur.fetchone()
        if not atualizacao:
            raise HTTPException(
                status_code=404, 
                detail="Atualização não encontrada ou já foi processada"
            )
        
        cadastro_id = atualizacao["cadastro_id"]
        dados_novos = atualizacao["dados_novos"]
        
        # Atualizar cadastro com novos dados (merge)
        if dados_novos:
            cur.execute("""
                UPDATE cadastros 
                SET dados = dados || %s::jsonb
                WHERE id = %s
            """, (json.dumps(dados_novos), cadastro_id))
        
        # Marcar atualização como aprovada
        cur.execute("""
            UPDATE atualizacoes_cadastrais 
            SET status = 'aprovada', 
                analisado_em = CURRENT_TIMESTAMP,
                analisado_por = %s,
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (admin["email"], atualizacao_id))
        
        conn.commit()
        
        # Notificar cliente
        cur.execute("""
            SELECT dados->>'nome' as nome, dados->>'email' as email 
            FROM cadastros WHERE id = %s
        """, (cadastro_id,))
        cliente = cur.fetchone()
        
        email_enviado = False
        if cliente and cliente["email"]:
            email_enviado = enviar_email_atualizacao_aprovada(cliente["email"], cliente["nome"] or "Cliente")
        
        cur.close()
        conn.close()
        
        return {
            "success": True, 
            "message": "Atualização aprovada e dados atualizados com sucesso",
            "email_enviado": email_enviado
        }
        
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Erro ao aprovar atualização: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/admin/atualizacoes/{atualizacao_id}/rejeitar")
async def rejeitar_atualizacao(
    atualizacao_id: int, 
    dados: RejeicaoAtualizacao,
    admin = Depends(verificar_admin)
):
    """Rejeita a atualização com um motivo."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT cadastro_id FROM atualizacoes_cadastrais 
            WHERE id = %s AND status = 'enviada'
        """, (atualizacao_id,))
        
        result = cur.fetchone()
        if not result:
            raise HTTPException(
                status_code=404, 
                detail="Atualização não encontrada ou já foi processada"
            )
        
        cadastro_id = result["cadastro_id"]
        
        cur.execute("""
            UPDATE atualizacoes_cadastrais 
            SET status = 'rejeitada', 
                analisado_em = CURRENT_TIMESTAMP,
                analisado_por = %s,
                motivo_rejeicao = %s,
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (admin["email"], dados.motivo, atualizacao_id))
        
        conn.commit()
        
        # Notificar cliente
        cur.execute("""
            SELECT dados->>'nome' as nome, dados->>'email' as email 
            FROM cadastros WHERE id = %s
        """, (cadastro_id,))
        cliente = cur.fetchone()
        
        email_enviado = False
        if cliente and cliente["email"]:
            email_enviado = enviar_email_atualizacao_rejeitada(
                cliente["email"], 
                cliente["nome"] or "Cliente", 
                dados.motivo
            )
        
        cur.close()
        conn.close()
        
        return {
            "success": True, 
            "message": "Atualização rejeitada",
            "email_enviado": email_enviado
        }
        
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Erro ao rejeitar atualização: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# ENDPOINTS - ATUALIZAÇÃO CADASTRAL (CLIENTE)
# ============================================

@app.get("/api/cliente/atualizacao-pendente")
async def verificar_atualizacao_pendente(cliente = Depends(verificar_token_cliente)):
    """Verifica se há solicitação de atualização pendente para o cliente."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, tipo, motivo_solicitacao, solicitado_em, status
            FROM atualizacoes_cadastrais 
            WHERE cadastro_id = %s AND status = 'pendente'
            ORDER BY criado_em DESC LIMIT 1
        """, (cliente["cadastro_id"],))
        
        result = cur.fetchone()
        cur.close()
        conn.close()
        
        if result:
            return {
                "tem_solicitacao": True,
                "solicitacao": {
                    "id": result["id"],
                    "tipo": result["tipo"],
                    "motivo": result["motivo_solicitacao"],
                    "data": result["solicitado_em"].isoformat() if result["solicitado_em"] else None,
                    "status": result["status"]
                }
            }
        
        return {"tem_solicitacao": False, "solicitacao": None}
        
    except Exception as e:
        logger.error(f"Erro ao verificar solicitação pendente: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/cliente/iniciar-atualizacao")
async def iniciar_atualizacao_espontanea(cliente = Depends(verificar_token_cliente)):
    """Cliente inicia uma atualização espontânea de seus dados."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT id, status FROM atualizacoes_cadastrais 
            WHERE cadastro_id = %s AND status IN ('pendente', 'enviada')
            ORDER BY criado_em DESC LIMIT 1
        """, (cliente["cadastro_id"],))
        
        existente = cur.fetchone()
        if existente:
            cur.close()
            conn.close()
            return {
                "success": True, 
                "id": existente["id"], 
                "status": existente["status"],
                "message": "Já existe uma atualização em andamento"
            }
        
        cur.execute("""
            INSERT INTO atualizacoes_cadastrais 
            (cadastro_id, tipo, status)
            VALUES (%s, 'espontanea', 'pendente')
            RETURNING id
        """, (cliente["cadastro_id"],))
        
        atualizacao_id = cur.fetchone()["id"]
        conn.commit()
        
        cur.close()
        conn.close()
        
        return {"success": True, "id": atualizacao_id, "message": "Atualização iniciada"}
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Erro ao iniciar atualização: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/cliente/enviar-atualizacao")
async def enviar_atualizacao_cadastral(
    dados: EnvioAtualizacao,
    cliente = Depends(verificar_token_cliente)
):
    """Cliente envia dados atualizados para análise do escritório."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        atualizacao_id = dados.atualizacao_id
        
        if atualizacao_id:
            cur.execute("""
                UPDATE atualizacoes_cadastrais 
                SET dados_novos = %s,
                    documentos_novos = %s,
                    status = 'enviada',
                    enviado_em = CURRENT_TIMESTAMP,
                    atualizado_em = CURRENT_TIMESTAMP
                WHERE id = %s AND cadastro_id = %s AND status = 'pendente'
                RETURNING id
            """, (
                json.dumps(dados.dados),
                json.dumps(dados.documentos),
                atualizacao_id,
                cliente["cadastro_id"]
            ))
            
            result = cur.fetchone()
            if not result:
                raise HTTPException(
                    status_code=400, 
                    detail="Solicitação não encontrada ou já foi enviada"
                )
        else:
            cur.execute("""
                INSERT INTO atualizacoes_cadastrais 
                (cadastro_id, tipo, status, dados_novos, documentos_novos, enviado_em)
                VALUES (%s, 'espontanea', 'enviada', %s, %s, CURRENT_TIMESTAMP)
                RETURNING id
            """, (
                cliente["cadastro_id"],
                json.dumps(dados.dados),
                json.dumps(dados.documentos)
            ))
            
            atualizacao_id = cur.fetchone()["id"]
        
        conn.commit()
        
        # Buscar nome do cliente para email
        cur.execute("""
            SELECT dados->>'nome' as nome FROM cadastros WHERE id = %s
        """, (cliente["cadastro_id"],))
        result = cur.fetchone()
        nome_cliente = result["nome"] if result else cliente["cadastro_id"]
        
        cur.close()
        conn.close()
        
        enviar_email_nova_atualizacao_admin(cliente["cadastro_id"], nome_cliente)
        
        return {
            "success": True, 
            "message": "Atualização enviada para análise do escritório",
            "id": atualizacao_id
        }
        
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Erro ao enviar atualização: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/cliente/minhas-atualizacoes")
async def listar_minhas_atualizacoes(cliente = Depends(verificar_token_cliente)):
    """Lista histórico de atualizações cadastrais do cliente logado."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, tipo, status, motivo_solicitacao, 
                   solicitado_em, enviado_em, analisado_em, motivo_rejeicao
            FROM atualizacoes_cadastrais 
            WHERE cadastro_id = %s 
            ORDER BY criado_em DESC
        """, (cliente["cadastro_id"],))
        
        atualizacoes = []
        for row in cur.fetchall():
            atualizacao = dict(row)
            for campo in ['solicitado_em', 'enviado_em', 'analisado_em']:
                if atualizacao.get(campo):
                    atualizacao[campo] = atualizacao[campo].isoformat()
            atualizacoes.append(atualizacao)
        
        cur.close()
        conn.close()
        
        return {"atualizacoes": atualizacoes}
        
    except Exception as e:
        logger.error(f"Erro ao listar atualizações: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# ENDPOINTS - BACKUP E GERENCIAMENTO DE DOCUMENTOS
# ============================================

@app.get("/api/admin/backup/listar-documentos")
async def admin_listar_todos_documentos(
    usuario: dict = Depends(verificar_admin)
):
    """
    Lista TODOS os documentos do sistema organizados por cliente.
    Retorna estrutura para seleção individual de arquivos.
    """
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Buscar todos os cadastros
        cur.execute("""
            SELECT id, dados, documentos, arquivos_gerados, documentos_assinados
            FROM cadastros ORDER BY data_hora DESC
        """)
        cadastros = cur.fetchall()
        
        resultado = []
        
        for cadastro in cadastros:
            cadastro_id = cadastro["id"]
            dados = cadastro["dados"] if isinstance(cadastro["dados"], dict) else json.loads(cadastro["dados"] or "{}")
            nome_cliente = dados.get("nome", "Sem nome")
            
            documentos_cliente = {
                "cadastro_id": cadastro_id,
                "nome_cliente": nome_cliente,
                "documentos": []
            }
            
            # 1. Documentos enviados pelo cliente no cadastro inicial
            # Os documentos podem ser strings (nome do arquivo) ou objetos com campo "arquivo"
            docs_cadastro = cadastro["documentos"] if isinstance(cadastro["documentos"], list) else json.loads(cadastro["documentos"] or "[]")
            for doc in docs_cadastro:
                arquivo = None
                nome = None

                if isinstance(doc, str):
                    # Formato: apenas nome do arquivo (string)
                    arquivo = doc
                    nome = doc
                elif isinstance(doc, dict) and doc.get("arquivo"):
                    # Formato: objeto com campo "arquivo"
                    arquivo = doc.get("arquivo")
                    nome = doc.get("nome", arquivo)

                if arquivo:
                    # Arquivos estão em UPLOADS_DIR/cadastro_id/filename
                    caminho = os.path.join(UPLOADS_DIR, cadastro_id, arquivo)
                    documentos_cliente["documentos"].append({
                        "id": f"cadastro_{cadastro_id}_{arquivo}",
                        "tipo": "cadastro_inicial",
                        "nome": nome,
                        "arquivo": arquivo,
                        "caminho": caminho,
                        "existe": os.path.exists(caminho),
                        "tamanho": os.path.getsize(caminho) if os.path.exists(caminho) else 0
                    })
            
            # 2. Documentos gerados (contrato e procuração)
            arq_gerados = cadastro["arquivos_gerados"] if isinstance(cadastro["arquivos_gerados"], dict) else json.loads(cadastro["arquivos_gerados"] or "{}")
            for tipo, caminho in arq_gerados.items():
                if caminho:
                    caminho_completo = caminho if os.path.isabs(caminho) else os.path.join(BASE_DIR, caminho)
                    documentos_cliente["documentos"].append({
                        "id": f"gerado_{cadastro_id}_{tipo}",
                        "tipo": "documento_gerado",
                        "nome": f"{tipo.replace('_', ' ').title()}",
                        "arquivo": os.path.basename(caminho),
                        "caminho": caminho_completo,
                        "existe": os.path.exists(caminho_completo),
                        "tamanho": os.path.getsize(caminho_completo) if os.path.exists(caminho_completo) else 0
                    })
            
            # 3. Documentos assinados
            # Podem ser strings (nome do arquivo) ou objetos
            docs_assinados = cadastro["documentos_assinados"] if isinstance(cadastro["documentos_assinados"], list) else json.loads(cadastro["documentos_assinados"] or "[]")
            for doc in docs_assinados:
                arquivo = None
                nome = None

                if isinstance(doc, str):
                    # Formato: apenas nome do arquivo (string)
                    arquivo = doc
                    nome = doc
                elif isinstance(doc, dict) and doc.get("arquivo"):
                    # Formato: objeto com campo "arquivo"
                    arquivo = doc.get("arquivo")
                    nome = doc.get("nome", arquivo)

                if arquivo:
                    # Arquivos assinados estão em UPLOADS_DIR/documentos_assinados/cadastro_id/filename
                    caminho = os.path.join(UPLOADS_DIR, "documentos_assinados", cadastro_id, arquivo)
                    documentos_cliente["documentos"].append({
                        "id": f"assinado_{cadastro_id}_{arquivo}",
                        "tipo": "documento_assinado",
                        "nome": nome,
                        "arquivo": arquivo,
                        "caminho": caminho,
                        "existe": os.path.exists(caminho),
                        "tamanho": os.path.getsize(caminho) if os.path.exists(caminho) else 0
                    })
            
            # 4. Documentos extras enviados pelo cliente (tabela documentos_extras)
            cur.execute("""
                SELECT id, nome_original, arquivo_path, descricao, criado_em
                FROM documentos_extras WHERE cadastro_id = %s
            """, (cadastro_id,))
            docs_extras = cur.fetchall()
            for doc in docs_extras:
                documentos_cliente["documentos"].append({
                    "id": f"extra_{doc['id']}",
                    "db_id": doc["id"],
                    "tipo": "documento_extra",
                    "nome": doc["nome_original"],
                    "descricao": doc.get("descricao", ""),
                    "caminho": doc["arquivo_path"],
                    "existe": os.path.exists(doc["arquivo_path"]),
                    "tamanho": os.path.getsize(doc["arquivo_path"]) if os.path.exists(doc["arquivo_path"]) else 0,
                    "data": doc["criado_em"].isoformat() if doc["criado_em"] else None
                })
            
            # 5. Documentos enviados pelo admin para o cliente (tabela documentos_admin)
            cur.execute("""
                SELECT id, nome_original, arquivo_path, descricao, criado_em
                FROM documentos_admin WHERE cadastro_id = %s
            """, (cadastro_id,))
            docs_admin = cur.fetchall()
            for doc in docs_admin:
                documentos_cliente["documentos"].append({
                    "id": f"admin_{doc['id']}",
                    "db_id": doc["id"],
                    "tipo": "documento_admin",
                    "nome": doc["nome_original"],
                    "descricao": doc.get("descricao", ""),
                    "caminho": doc["arquivo_path"],
                    "existe": os.path.exists(doc["arquivo_path"]),
                    "tamanho": os.path.getsize(doc["arquivo_path"]) if os.path.exists(doc["arquivo_path"]) else 0,
                    "data": doc["criado_em"].isoformat() if doc["criado_em"] else None
                })
            
            # 6. Comprovantes de pagamento
            cur.execute("""
                SELECT c.id, c.arquivo_nome, c.arquivo_path, c.enviado_em, c.status,
                       p.numero as parcela_numero, ch.descricao as contrato_descricao
                FROM comprovantes c
                JOIN parcelas p ON c.parcela_id = p.id
                JOIN contratos_honorarios ch ON p.contrato_id = ch.id
                WHERE ch.cadastro_id = %s
            """, (cadastro_id,))
            comprovantes = cur.fetchall()
            for comp in comprovantes:
                documentos_cliente["documentos"].append({
                    "id": f"comprovante_{comp['id']}",
                    "db_id": comp["id"],
                    "tipo": "comprovante",
                    "nome": f"Comprovante - Parcela {comp['parcela_numero']} - {comp.get('contrato_descricao', '')}",
                    "arquivo": comp["arquivo_nome"],
                    "caminho": comp["arquivo_path"],
                    "existe": os.path.exists(comp["arquivo_path"]) if comp["arquivo_path"] else False,
                    "tamanho": os.path.getsize(comp["arquivo_path"]) if comp["arquivo_path"] and os.path.exists(comp["arquivo_path"]) else 0,
                    "status": comp["status"],
                    "data": comp["enviado_em"].isoformat() if comp["enviado_em"] else None
                })

            # 7. Documentos da Demanda (tabela documentos_demanda)
            cur.execute("""
                SELECT id, tipo_documento, nome_arquivo, nome_original, arquivo_path, descricao, criado_em
                FROM documentos_demanda WHERE cadastro_id = %s
            """, (cadastro_id,))
            docs_demanda = cur.fetchall()
            for doc in docs_demanda:
                documentos_cliente["documentos"].append({
                    "id": f"demanda_{doc['id']}",
                    "db_id": doc["id"],
                    "tipo": "documento_demanda",
                    "nome": doc["nome_original"],
                    "tipo_documento": doc["tipo_documento"],
                    "descricao": doc.get("descricao", ""),
                    "caminho": doc["arquivo_path"],
                    "existe": os.path.exists(doc["arquivo_path"]) if doc["arquivo_path"] else False,
                    "tamanho": os.path.getsize(doc["arquivo_path"]) if doc["arquivo_path"] and os.path.exists(doc["arquivo_path"]) else 0,
                    "data": doc["criado_em"].isoformat() if doc["criado_em"] else None
                })

            # Calcular totais
            docs_existentes = [d for d in documentos_cliente["documentos"] if d.get("existe")]
            documentos_cliente["total_documentos"] = len(documentos_cliente["documentos"])
            documentos_cliente["documentos_existentes"] = len(docs_existentes)
            documentos_cliente["tamanho_total"] = sum(d.get("tamanho", 0) for d in docs_existentes)
            
            resultado.append(documentos_cliente)
        
        cur.close()
        conn.close()
        
        # Calcular totais gerais
        total_geral = sum(c["total_documentos"] for c in resultado)
        existentes_geral = sum(c["documentos_existentes"] for c in resultado)
        tamanho_geral = sum(c["tamanho_total"] for c in resultado)
        
        return {
            "clientes": resultado,
            "resumo": {
                "total_clientes": len(resultado),
                "total_documentos": total_geral,
                "documentos_existentes": existentes_geral,
                "tamanho_total": tamanho_geral,
                "tamanho_formatado": formatar_tamanho(tamanho_geral)
            }
        }
    except Exception as e:
        logger.error(f"Erro ao listar documentos para backup: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def formatar_tamanho(bytes_size: int) -> str:
    """Formata tamanho em bytes para formato legível."""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if bytes_size < 1024:
            return f"{bytes_size:.1f} {unit}"
        bytes_size /= 1024
    return f"{bytes_size:.1f} TB"


class BackupRequest(BaseModel):
    documentos_ids: List[str]  # Lista de IDs dos documentos selecionados
    incluir_dados_json: bool = True  # Se inclui JSON com dados dos clientes


@app.post("/api/admin/backup/download")
async def admin_download_backup(
    request: BackupRequest,
    usuario: dict = Depends(verificar_admin)
):
    """
    Gera um ZIP com os documentos selecionados.
    Organiza por cliente em pastas separadas.
    """
    if not request.documentos_ids:
        raise HTTPException(status_code=400, detail="Nenhum documento selecionado")
    
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Criar arquivo ZIP em memória
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            
            # Buscar todos os cadastros para organização
            cur.execute("SELECT id, dados FROM cadastros")
            cadastros = {row["id"]: row for row in cur.fetchall()}
            
            arquivos_adicionados = set()
            dados_clientes = {}
            
            for doc_id in request.documentos_ids:
                partes = doc_id.split("_", 2)
                if len(partes) < 2:
                    continue
                
                tipo = partes[0]
                
                # Identificar cadastro_id e buscar arquivo
                caminho_arquivo = None
                nome_arquivo = None
                cadastro_id = None
                
                if tipo == "cadastro":
                    # Formato: cadastro_{cadastro_id}_{arquivo}
                    cadastro_id = partes[1]
                    arquivo = "_".join(partes[2:]) if len(partes) > 2 else ""
                    caminho_arquivo = os.path.join(UPLOADS_DIR, arquivo)
                    nome_arquivo = arquivo
                    
                elif tipo == "gerado":
                    # Formato: gerado_{cadastro_id}_{tipo_doc}
                    cadastro_id = partes[1]
                    tipo_doc = "_".join(partes[2:]) if len(partes) > 2 else ""
                    cadastro = buscar_cadastro(cadastro_id)
                    if cadastro and cadastro.get("arquivos_gerados"):
                        caminho = cadastro["arquivos_gerados"].get(tipo_doc)
                        if caminho:
                            caminho_arquivo = caminho if os.path.isabs(caminho) else os.path.join(BASE_DIR, caminho)
                            nome_arquivo = os.path.basename(caminho)
                
                elif tipo == "assinado":
                    # Formato: assinado_{cadastro_id}_{arquivo}
                    cadastro_id = partes[1]
                    arquivo = "_".join(partes[2:]) if len(partes) > 2 else ""
                    caminho_arquivo = os.path.join(UPLOADS_DIR, arquivo)
                    nome_arquivo = arquivo
                
                elif tipo == "extra":
                    # Formato: extra_{db_id}
                    db_id = int(partes[1])
                    cur.execute("SELECT * FROM documentos_extras WHERE id = %s", (db_id,))
                    doc = cur.fetchone()
                    if doc:
                        cadastro_id = doc["cadastro_id"]
                        caminho_arquivo = doc["arquivo_path"]
                        nome_arquivo = doc["nome_original"]
                
                elif tipo == "admin":
                    # Formato: admin_{db_id}
                    db_id = int(partes[1])
                    cur.execute("SELECT * FROM documentos_admin WHERE id = %s", (db_id,))
                    doc = cur.fetchone()
                    if doc:
                        cadastro_id = doc["cadastro_id"]
                        caminho_arquivo = doc["arquivo_path"]
                        nome_arquivo = doc["nome_original"]
                
                elif tipo == "comprovante":
                    # Formato: comprovante_{db_id}
                    db_id = int(partes[1])
                    cur.execute("""
                        SELECT c.*, ch.cadastro_id 
                        FROM comprovantes c
                        JOIN parcelas p ON c.parcela_id = p.id
                        JOIN contratos_honorarios ch ON p.contrato_id = ch.id
                        WHERE c.id = %s
                    """, (db_id,))
                    doc = cur.fetchone()
                    if doc:
                        cadastro_id = doc["cadastro_id"]
                        caminho_arquivo = doc["arquivo_path"]
                        nome_arquivo = doc["arquivo_nome"]
                
                # Adicionar arquivo ao ZIP
                if caminho_arquivo and os.path.exists(caminho_arquivo) and caminho_arquivo not in arquivos_adicionados:
                    # Buscar nome do cliente
                    nome_cliente = "Sem_Nome"
                    if cadastro_id and cadastro_id in cadastros:
                        dados = cadastros[cadastro_id]["dados"]
                        if isinstance(dados, str):
                            dados = json.loads(dados)
                        nome_cliente = dados.get("nome", "Sem_Nome").replace(" ", "_").replace("/", "-")[:50]
                    
                    # Criar pasta do cliente
                    pasta_cliente = f"{cadastro_id}_{nome_cliente}"
                    
                    # Adicionar ao ZIP
                    zip_file.write(caminho_arquivo, f"{pasta_cliente}/{nome_arquivo}")
                    arquivos_adicionados.add(caminho_arquivo)
                    
                    # Guardar dados do cliente para JSON
                    if cadastro_id not in dados_clientes:
                        dados_clientes[cadastro_id] = {
                            "cadastro_id": cadastro_id,
                            "nome": nome_cliente,
                            "arquivos": []
                        }
                    dados_clientes[cadastro_id]["arquivos"].append(nome_arquivo)
            
            # Incluir JSON com dados dos clientes
            if request.incluir_dados_json:
                # Buscar dados completos dos clientes incluídos
                for cadastro_id in dados_clientes.keys():
                    cadastro = buscar_cadastro(cadastro_id)
                    if cadastro:
                        dados_clientes[cadastro_id]["dados_cadastro"] = cadastro["dados"]
                        dados_clientes[cadastro_id]["status"] = cadastro["status"]
                        dados_clientes[cadastro_id]["data_cadastro"] = cadastro["data"]
                        
                        # Buscar processos
                        processos = listar_processos(cadastro_id)
                        dados_clientes[cadastro_id]["processos"] = processos
                        
                        # Buscar contratos
                        contratos = listar_contratos(cadastro_id)
                        dados_clientes[cadastro_id]["contratos"] = contratos
                
                # Adicionar JSON ao ZIP
                json_content = json.dumps(list(dados_clientes.values()), indent=2, ensure_ascii=False, default=str)
                zip_file.writestr("dados_clientes.json", json_content)
        
        cur.close()
        conn.close()
        
        # Retornar o ZIP
        zip_buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"backup_vaucher_alvares_{timestamp}.zip"
        
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Erro ao gerar backup: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class DeleteDocumentosRequest(BaseModel):
    documentos_ids: List[str]  # Lista de IDs dos documentos a deletar
    confirmar: bool = False  # Confirmação obrigatória


@app.post("/api/admin/backup/deletar-documentos")
async def admin_deletar_documentos_selecionados(
    request: DeleteDocumentosRequest,
    usuario: dict = Depends(verificar_admin)
):
    """
    Deleta múltiplos documentos selecionados.
    Requer confirmação explícita.
    """
    if not request.confirmar:
        raise HTTPException(
            status_code=400, 
            detail="Confirmação obrigatória. Defina 'confirmar: true' para prosseguir."
        )
    
    if not request.documentos_ids:
        raise HTTPException(status_code=400, detail="Nenhum documento selecionado")
    
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        deletados = []
        erros = []
        
        for doc_id in request.documentos_ids:
            try:
                partes = doc_id.split("_", 2)
                if len(partes) < 2:
                    erros.append({"id": doc_id, "erro": "ID inválido"})
                    continue
                
                tipo = partes[0]
                sucesso = False
                
                if tipo == "cadastro":
                    # Documentos do cadastro inicial - remover do JSON
                    cadastro_id = partes[1]
                    arquivo = "_".join(partes[2:]) if len(partes) > 2 else ""
                    # Arquivos do cliente estão em UPLOADS_DIR/cadastro_id/filename
                    caminho = os.path.join(UPLOADS_DIR, cadastro_id, arquivo)

                    # Remover arquivo físico
                    if os.path.exists(caminho):
                        os.remove(caminho)

                    # Atualizar JSON no banco
                    cadastro = buscar_cadastro(cadastro_id)
                    if cadastro:
                        docs = cadastro.get("documentos", [])
                        # Filtrar - docs podem ser strings ou dicts
                        docs_atualizados = []
                        for d in docs:
                            if isinstance(d, str):
                                if d != arquivo:
                                    docs_atualizados.append(d)
                            elif isinstance(d, dict):
                                if d.get("arquivo") != arquivo:
                                    docs_atualizados.append(d)
                        cur.execute(
                            "UPDATE cadastros SET documentos = %s WHERE id = %s",
                            (json.dumps(docs_atualizados), cadastro_id)
                        )
                        conn.commit()
                    sucesso = True
                    
                elif tipo == "gerado":
                    # Documentos gerados - remover arquivo e atualizar JSON
                    cadastro_id = partes[1]
                    tipo_doc = "_".join(partes[2:]) if len(partes) > 2 else ""
                    
                    cadastro = buscar_cadastro(cadastro_id)
                    if cadastro and cadastro.get("arquivos_gerados"):
                        caminho = cadastro["arquivos_gerados"].get(tipo_doc)
                        if caminho:
                            caminho_completo = caminho if os.path.isabs(caminho) else os.path.join(BASE_DIR, caminho)
                            if os.path.exists(caminho_completo):
                                os.remove(caminho_completo)
                        
                        # Atualizar JSON
                        arq_gerados = cadastro["arquivos_gerados"]
                        arq_gerados[tipo_doc] = None
                        cur.execute(
                            "UPDATE cadastros SET arquivos_gerados = %s WHERE id = %s",
                            (json.dumps(arq_gerados), cadastro_id)
                        )
                        conn.commit()
                    sucesso = True
                    
                elif tipo == "assinado":
                    # Documentos assinados
                    cadastro_id = partes[1]
                    arquivo = "_".join(partes[2:]) if len(partes) > 2 else ""
                    # Arquivos assinados estão em UPLOADS_DIR/documentos_assinados/cadastro_id/filename
                    caminho = os.path.join(UPLOADS_DIR, "documentos_assinados", cadastro_id, arquivo)

                    if os.path.exists(caminho):
                        os.remove(caminho)

                    cadastro = buscar_cadastro(cadastro_id)
                    if cadastro:
                        docs = cadastro.get("documentos_assinados", [])
                        # Filtrar - docs podem ser strings ou dicts
                        docs_atualizados = []
                        for d in docs:
                            if isinstance(d, str):
                                if d != arquivo:
                                    docs_atualizados.append(d)
                            elif isinstance(d, dict):
                                if d.get("arquivo") != arquivo:
                                    docs_atualizados.append(d)
                        cur.execute(
                            "UPDATE cadastros SET documentos_assinados = %s WHERE id = %s",
                            (json.dumps(docs_atualizados), cadastro_id)
                        )
                        conn.commit()
                    sucesso = True
                    
                elif tipo == "extra":
                    # Documentos extras - deletar do banco e arquivo
                    db_id = int(partes[1])
                    cur.execute("SELECT arquivo_path FROM documentos_extras WHERE id = %s", (db_id,))
                    doc = cur.fetchone()
                    if doc and doc["arquivo_path"] and os.path.exists(doc["arquivo_path"]):
                        os.remove(doc["arquivo_path"])
                    cur.execute("DELETE FROM documentos_extras WHERE id = %s", (db_id,))
                    conn.commit()
                    sucesso = True
                    
                elif tipo == "admin":
                    # Documentos admin
                    db_id = int(partes[1])
                    cur.execute("SELECT arquivo_path FROM documentos_admin WHERE id = %s", (db_id,))
                    doc = cur.fetchone()
                    if doc and doc["arquivo_path"] and os.path.exists(doc["arquivo_path"]):
                        os.remove(doc["arquivo_path"])
                    cur.execute("DELETE FROM documentos_admin WHERE id = %s", (db_id,))
                    conn.commit()
                    sucesso = True
                    
                elif tipo == "comprovante":
                    # Comprovantes
                    db_id = int(partes[1])
                    cur.execute("SELECT arquivo_path FROM comprovantes WHERE id = %s", (db_id,))
                    doc = cur.fetchone()
                    if doc and doc["arquivo_path"] and os.path.exists(doc["arquivo_path"]):
                        os.remove(doc["arquivo_path"])
                    cur.execute("DELETE FROM comprovantes WHERE id = %s", (db_id,))
                    conn.commit()
                    sucesso = True

                elif tipo == "demanda":
                    # Documentos da demanda
                    db_id = int(partes[1])
                    cur.execute("SELECT arquivo_path FROM documentos_demanda WHERE id = %s", (db_id,))
                    doc = cur.fetchone()
                    if doc and doc["arquivo_path"] and os.path.exists(doc["arquivo_path"]):
                        os.remove(doc["arquivo_path"])
                    cur.execute("DELETE FROM documentos_demanda WHERE id = %s", (db_id,))
                    conn.commit()
                    sucesso = True

                if sucesso:
                    deletados.append(doc_id)
                else:
                    erros.append({"id": doc_id, "erro": "Tipo não reconhecido"})
                    
            except Exception as e:
                erros.append({"id": doc_id, "erro": str(e)})
        
        cur.close()
        conn.close()
        
        return {
            "success": True,
            "deletados": len(deletados),
            "erros": len(erros),
            "detalhes": {
                "ids_deletados": deletados,
                "erros": erros
            }
        }
    except Exception as e:
        logger.error(f"Erro ao deletar documentos: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/backup/download-documento")
async def admin_download_documento_backup(
    doc_id: str,
    usuario: dict = Depends(verificar_admin)
):
    """
    Faz download de um documento específico pelo seu ID de backup.
    """
    try:
        partes = doc_id.split("_", 2)
        if len(partes) < 2:
            raise HTTPException(status_code=400, detail="ID de documento inválido")

        tipo = partes[0]
        caminho = None
        nome_arquivo = None

        if tipo == "cadastro":
            cadastro_id = partes[1]
            arquivo = "_".join(partes[2:]) if len(partes) > 2 else ""
            # Arquivos do cliente estão em UPLOADS_DIR/cadastro_id/filename
            caminho = os.path.join(UPLOADS_DIR, cadastro_id, arquivo)
            nome_arquivo = arquivo

        elif tipo == "gerado":
            cadastro_id = partes[1]
            tipo_doc = "_".join(partes[2:]) if len(partes) > 2 else ""
            cadastro = buscar_cadastro(cadastro_id)
            if cadastro and cadastro.get("arquivos_gerados"):
                caminho_rel = cadastro["arquivos_gerados"].get(tipo_doc)
                if caminho_rel:
                    caminho = caminho_rel if os.path.isabs(caminho_rel) else os.path.join(BASE_DIR, caminho_rel)
                    nome_arquivo = os.path.basename(caminho)

        elif tipo == "assinado":
            cadastro_id = partes[1]
            arquivo = "_".join(partes[2:]) if len(partes) > 2 else ""
            # Arquivos assinados estão em UPLOADS_DIR/documentos_assinados/cadastro_id/filename
            caminho = os.path.join(UPLOADS_DIR, "documentos_assinados", cadastro_id, arquivo)
            nome_arquivo = arquivo

        elif tipo == "extra":
            db_id = int(partes[1])
            conn = get_db()
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute("SELECT arquivo_path, nome_original FROM documentos_extras WHERE id = %s", (db_id,))
            doc = cur.fetchone()
            cur.close()
            conn.close()
            if doc:
                caminho = doc["arquivo_path"]
                nome_arquivo = doc["nome_original"]

        elif tipo == "admin":
            db_id = int(partes[1])
            conn = get_db()
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute("SELECT arquivo_path, nome_original FROM documentos_admin WHERE id = %s", (db_id,))
            doc = cur.fetchone()
            cur.close()
            conn.close()
            if doc:
                caminho = doc["arquivo_path"]
                nome_arquivo = doc["nome_original"]

        elif tipo == "comprovante":
            db_id = int(partes[1])
            conn = get_db()
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute("SELECT arquivo_path, arquivo_nome FROM comprovantes WHERE id = %s", (db_id,))
            doc = cur.fetchone()
            cur.close()
            conn.close()
            if doc:
                caminho = doc["arquivo_path"]
                nome_arquivo = doc["arquivo_nome"]

        elif tipo == "demanda":
            db_id = int(partes[1])
            conn = get_db()
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute("SELECT arquivo_path, nome_original FROM documentos_demanda WHERE id = %s", (db_id,))
            doc = cur.fetchone()
            cur.close()
            conn.close()
            if doc:
                caminho = doc["arquivo_path"]
                nome_arquivo = doc["nome_original"]

        if not caminho or not os.path.exists(caminho):
            raise HTTPException(status_code=404, detail="Arquivo não encontrado")

        return FileResponse(
            path=caminho,
            filename=nome_arquivo or os.path.basename(caminho),
            media_type="application/octet-stream"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao baixar documento: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class DownloadSelecionadosRequest(BaseModel):
    documentos_ids: list[str]


@app.post("/api/admin/backup/download-selecionados")
async def admin_download_documentos_selecionados(
    request: DownloadSelecionadosRequest,
    usuario: dict = Depends(verificar_admin)
):
    """
    Faz download de múltiplos documentos selecionados em um arquivo ZIP.
    """
    if not request.documentos_ids:
        raise HTTPException(status_code=400, detail="Nenhum documento selecionado")

    try:
        # Criar ZIP em memória
        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            arquivos_adicionados = 0
            nomes_usados = {}  # Para evitar nomes duplicados

            for doc_id in request.documentos_ids:
                try:
                    partes = doc_id.split("_", 2)
                    if len(partes) < 2:
                        continue

                    tipo = partes[0]
                    caminho = None
                    nome_arquivo = None
                    pasta = ""

                    if tipo == "cadastro":
                        cadastro_id = partes[1]
                        arquivo = "_".join(partes[2:]) if len(partes) > 2 else ""
                        caminho = os.path.join(UPLOADS_DIR, cadastro_id, arquivo)
                        nome_arquivo = arquivo
                        pasta = f"{cadastro_id}/documentos/"

                    elif tipo == "gerado":
                        cadastro_id = partes[1]
                        tipo_doc = "_".join(partes[2:]) if len(partes) > 2 else ""
                        cadastro = buscar_cadastro(cadastro_id)
                        if cadastro and cadastro.get("arquivos_gerados"):
                            caminho_rel = cadastro["arquivos_gerados"].get(tipo_doc)
                            if caminho_rel:
                                caminho = caminho_rel if os.path.isabs(caminho_rel) else os.path.join(BASE_DIR, caminho_rel)
                                nome_arquivo = os.path.basename(caminho)
                                pasta = f"{cadastro_id}/gerados/"

                    elif tipo == "assinado":
                        cadastro_id = partes[1]
                        arquivo = "_".join(partes[2:]) if len(partes) > 2 else ""
                        caminho = os.path.join(UPLOADS_DIR, "documentos_assinados", cadastro_id, arquivo)
                        nome_arquivo = arquivo
                        pasta = f"documentos_assinados/{cadastro_id}/"

                    elif tipo == "extra":
                        db_id = int(partes[1])
                        conn = get_db()
                        cur = conn.cursor(cursor_factory=RealDictCursor)
                        cur.execute("SELECT arquivo_path, nome_original, cadastro_id FROM documentos_extras WHERE id = %s", (db_id,))
                        doc = cur.fetchone()
                        cur.close()
                        conn.close()
                        if doc:
                            caminho = doc["arquivo_path"]
                            nome_arquivo = doc["nome_original"]
                            pasta = f"{doc['cadastro_id']}/extras/"

                    elif tipo == "admin":
                        db_id = int(partes[1])
                        conn = get_db()
                        cur = conn.cursor(cursor_factory=RealDictCursor)
                        cur.execute("SELECT arquivo_path, nome_original, cadastro_id FROM documentos_admin WHERE id = %s", (db_id,))
                        doc = cur.fetchone()
                        cur.close()
                        conn.close()
                        if doc:
                            caminho = doc["arquivo_path"]
                            nome_arquivo = doc["nome_original"]
                            pasta = f"{doc['cadastro_id']}/admin/"

                    elif tipo == "comprovante":
                        db_id = int(partes[1])
                        conn = get_db()
                        cur = conn.cursor(cursor_factory=RealDictCursor)
                        cur.execute("""
                            SELECT c.arquivo_path, c.arquivo_nome, ch.cadastro_id
                            FROM comprovantes c
                            JOIN parcelas p ON c.parcela_id = p.id
                            JOIN contratos_honorarios ch ON p.contrato_id = ch.id
                            WHERE c.id = %s
                        """, (db_id,))
                        doc = cur.fetchone()
                        cur.close()
                        conn.close()
                        if doc:
                            caminho = doc["arquivo_path"]
                            nome_arquivo = doc["arquivo_nome"]
                            pasta = f"{doc['cadastro_id']}/comprovantes/"

                    elif tipo == "demanda":
                        db_id = int(partes[1])
                        conn = get_db()
                        cur = conn.cursor(cursor_factory=RealDictCursor)
                        cur.execute("SELECT arquivo_path, nome_original, cadastro_id FROM documentos_demanda WHERE id = %s", (db_id,))
                        doc = cur.fetchone()
                        cur.close()
                        conn.close()
                        if doc:
                            caminho = doc["arquivo_path"]
                            nome_arquivo = doc["nome_original"]
                            pasta = f"{doc['cadastro_id']}/demanda/"

                    # Adicionar ao ZIP se arquivo existe
                    if caminho and os.path.exists(caminho) and nome_arquivo:
                        # Garantir nome único
                        nome_no_zip = f"{pasta}{nome_arquivo}"
                        if nome_no_zip in nomes_usados:
                            nomes_usados[nome_no_zip] += 1
                            base, ext = os.path.splitext(nome_arquivo)
                            nome_arquivo = f"{base}_{nomes_usados[nome_no_zip]}{ext}"
                            nome_no_zip = f"{pasta}{nome_arquivo}"
                        else:
                            nomes_usados[nome_no_zip] = 0

                        zip_file.write(caminho, nome_no_zip)
                        arquivos_adicionados += 1

                except Exception as e:
                    logger.error(f"Erro ao processar documento {doc_id}: {e}")
                    continue

        if arquivos_adicionados == 0:
            raise HTTPException(status_code=404, detail="Nenhum arquivo encontrado para download")

        # Preparar resposta
        zip_buffer.seek(0)
        from datetime import datetime
        data_atual = datetime.now().strftime("%Y%m%d_%H%M%S")

        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="backup_selecionados_{data_atual}.zip"'
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao criar ZIP de documentos: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/backup/download-completo")
async def admin_download_backup_completo(
    usuario: dict = Depends(verificar_admin)
):
    """
    Gera um backup COMPLETO de todos os clientes e documentos.
    Inclui JSON com todos os dados.
    """
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Criar arquivo ZIP em memória
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            
            # Buscar todos os cadastros
            cur.execute("SELECT * FROM cadastros ORDER BY data_hora DESC")
            cadastros = cur.fetchall()
            
            todos_dados = []
            
            for cadastro in cadastros:
                cadastro_id = cadastro["id"]
                dados = cadastro["dados"] if isinstance(cadastro["dados"], dict) else json.loads(cadastro["dados"] or "{}")
                nome_cliente = dados.get("nome", "Sem_Nome").replace(" ", "_").replace("/", "-")[:50]
                pasta_cliente = f"{cadastro_id}_{nome_cliente}"
                
                cliente_dados = {
                    "cadastro_id": cadastro_id,
                    "nome": dados.get("nome"),
                    "dados_cadastro": dados,
                    "status": cadastro["status"],
                    "data_cadastro": str(cadastro["data"]),
                    "arquivos_incluidos": []
                }
                
                # 1. Documentos do cadastro
                docs = cadastro["documentos"] if isinstance(cadastro["documentos"], list) else json.loads(cadastro["documentos"] or "[]")
                for doc in docs:
                    if isinstance(doc, dict) and doc.get("arquivo"):
                        caminho = os.path.join(UPLOADS_DIR, doc["arquivo"])
                        if os.path.exists(caminho):
                            zip_file.write(caminho, f"{pasta_cliente}/cadastro/{doc['arquivo']}")
                            cliente_dados["arquivos_incluidos"].append(f"cadastro/{doc['arquivo']}")
                
                # 2. Documentos gerados
                arq = cadastro["arquivos_gerados"] if isinstance(cadastro["arquivos_gerados"], dict) else json.loads(cadastro["arquivos_gerados"] or "{}")
                for tipo, caminho in arq.items():
                    if caminho:
                        caminho_completo = caminho if os.path.isabs(caminho) else os.path.join(BASE_DIR, caminho)
                        if os.path.exists(caminho_completo):
                            nome = os.path.basename(caminho)
                            zip_file.write(caminho_completo, f"{pasta_cliente}/gerados/{nome}")
                            cliente_dados["arquivos_incluidos"].append(f"gerados/{nome}")
                
                # 3. Documentos assinados
                docs_ass = cadastro["documentos_assinados"] if isinstance(cadastro["documentos_assinados"], list) else json.loads(cadastro["documentos_assinados"] or "[]")
                for doc in docs_ass:
                    if isinstance(doc, dict) and doc.get("arquivo"):
                        caminho = os.path.join(UPLOADS_DIR, doc["arquivo"])
                        if os.path.exists(caminho):
                            zip_file.write(caminho, f"{pasta_cliente}/assinados/{doc['arquivo']}")
                            cliente_dados["arquivos_incluidos"].append(f"assinados/{doc['arquivo']}")
                
                # 4. Documentos extras
                cur.execute("SELECT * FROM documentos_extras WHERE cadastro_id = %s", (cadastro_id,))
                for doc in cur.fetchall():
                    if doc["arquivo_path"] and os.path.exists(doc["arquivo_path"]):
                        zip_file.write(doc["arquivo_path"], f"{pasta_cliente}/extras/{doc['nome_original']}")
                        cliente_dados["arquivos_incluidos"].append(f"extras/{doc['nome_original']}")
                
                # 5. Documentos admin
                cur.execute("SELECT * FROM documentos_admin WHERE cadastro_id = %s", (cadastro_id,))
                for doc in cur.fetchall():
                    if doc["arquivo_path"] and os.path.exists(doc["arquivo_path"]):
                        zip_file.write(doc["arquivo_path"], f"{pasta_cliente}/documentos_escritorio/{doc['nome_original']}")
                        cliente_dados["arquivos_incluidos"].append(f"documentos_escritorio/{doc['nome_original']}")
                
                # 6. Comprovantes
                cur.execute("""
                    SELECT c.* FROM comprovantes c
                    JOIN parcelas p ON c.parcela_id = p.id
                    JOIN contratos_honorarios ch ON p.contrato_id = ch.id
                    WHERE ch.cadastro_id = %s
                """, (cadastro_id,))
                for comp in cur.fetchall():
                    if comp["arquivo_path"] and os.path.exists(comp["arquivo_path"]):
                        zip_file.write(comp["arquivo_path"], f"{pasta_cliente}/comprovantes/{comp['arquivo_nome']}")
                        cliente_dados["arquivos_incluidos"].append(f"comprovantes/{comp['arquivo_nome']}")
                
                # Buscar processos e contratos
                cliente_dados["processos"] = listar_processos(cadastro_id)
                cliente_dados["contratos"] = listar_contratos(cadastro_id)
                
                # Buscar financeiro
                financeiro = buscar_financeiro(cadastro_id)
                if financeiro:
                    cliente_dados["financeiro"] = financeiro
                
                todos_dados.append(cliente_dados)
            
            # Adicionar JSON completo
            json_content = json.dumps(todos_dados, indent=2, ensure_ascii=False, default=str)
            zip_file.writestr("backup_completo.json", json_content)
        
        cur.close()
        conn.close()
        
        zip_buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"backup_completo_vaucher_alvares_{timestamp}.zip"
        
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Erro ao gerar backup completo: {e}")
        raise HTTPException(status_code=500, detail=str(e))
        
# ============================================
# TERMOS DE USO E POLÍTICA DE PRIVACIDADE
# ============================================

@app.get("/api/termos/{tipo}")
async def get_termos_vigentes(tipo: str):
    """
    Retorna a versão vigente dos termos.
    Tipos: 'termos_uso' ou 'politica_privacidade'
    """
    if tipo not in ['termos_uso', 'politica_privacidade']:
        raise HTTPException(status_code=400, detail="Tipo inválido. Use 'termos_uso' ou 'politica_privacidade'")
    
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão com banco de dados")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, versao, conteudo, data_vigencia
            FROM termos_versoes
            WHERE tipo = %s AND ativo = TRUE
            ORDER BY data_vigencia DESC
            LIMIT 1
        """, (tipo,))
        
        row = cur.fetchone()
        cur.close()
        conn.close()
        
        if not row:
            raise HTTPException(status_code=404, detail="Termos não encontrados. Configure os termos no banco de dados.")
        
        return {
            "id": row["id"],
            "versao": row["versao"],
            "conteudo": row["conteudo"],
            "data_vigencia": row["data_vigencia"].isoformat() if row["data_vigencia"] else None
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao buscar termos: {e}")
        raise HTTPException(status_code=500, detail="Erro ao buscar termos")


@app.post("/api/cliente/aceitar-termos")
async def cliente_aceitar_termos(request):
    """
    Registra o aceite dos termos pelo cliente no momento do cadastro.
    Armazena IP, User-Agent e metadados para validade legal.
    Nota: Diferente de /api/aceitar-termos que é para usuários do painel admin.
    """
    from starlette.requests import Request
    
    # Obter dados do body
    try:
        data = await request.json()
    except:
        raise HTTPException(status_code=400, detail="JSON inválido")
    
    cadastro_id = data.get('cadastro_id')
    termos_uso_versao_id = data.get('termos_uso_versao_id')
    privacidade_versao_id = data.get('privacidade_versao_id')
    
    if not cadastro_id:
        raise HTTPException(status_code=400, detail="cadastro_id é obrigatório")
    
    if not termos_uso_versao_id and not privacidade_versao_id:
        raise HTTPException(status_code=400, detail="Pelo menos um ID de versão de termos é obrigatório")
    
    # Capturar IP do cliente
    ip_address = request.client.host if request.client else "0.0.0.0"
    
    # Tentar pegar IP real se estiver atrás de proxy/load balancer
    forwarded_for = request.headers.get('X-Forwarded-For')
    if forwarded_for:
        ip_address = forwarded_for.split(',')[0].strip()
    
    # Capturar User-Agent
    user_agent = request.headers.get('User-Agent', 'Não informado')
    
    # Metadados adicionais para prova legal
    metadados = {
        "accept_language": request.headers.get('Accept-Language'),
        "origin": request.headers.get('Origin'),
        "referer": request.headers.get('Referer'),
        "timestamp_servidor": datetime.now().isoformat(),
        "x_real_ip": request.headers.get('X-Real-IP'),
    }
    
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão com banco de dados")
    
    try:
        cur = conn.cursor()
        
        # Registrar aceite dos Termos de Uso
        if termos_uso_versao_id:
            cur.execute("""
                INSERT INTO aceites_termos (cadastro_id, termos_versao_id, ip_address, user_agent, metadados)
                VALUES (%s, %s, %s, %s, %s)
            """, (cadastro_id, termos_uso_versao_id, ip_address, user_agent, json.dumps(metadados)))
        
        # Registrar aceite da Política de Privacidade
        if privacidade_versao_id:
            cur.execute("""
                INSERT INTO aceites_termos (cadastro_id, termos_versao_id, ip_address, user_agent, metadados)
                VALUES (%s, %s, %s, %s, %s)
            """, (cadastro_id, privacidade_versao_id, ip_address, user_agent, json.dumps(metadados)))
        
        # Atualizar cadastro marcando que aceitou os termos
        cur.execute("""
            UPDATE cadastros 
            SET termos_aceitos = TRUE, termos_aceitos_em = NOW()
            WHERE id = %s
        """, (cadastro_id,))
        
        conn.commit()
        cur.close()
        conn.close()
        
        logger.info(f"Termos aceitos - Cadastro: {cadastro_id}, IP: {ip_address}")
        
        return {
            "success": True, 
            "message": "Termos aceitos e registrados com sucesso",
            "cadastro_id": cadastro_id,
            "ip_registrado": ip_address,
            "data_aceite": datetime.now().isoformat()
        }
        
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Erro ao registrar aceite de termos: {e}")
        raise HTTPException(status_code=500, detail="Erro ao registrar aceite dos termos")


@app.get("/api/admin/clientes/{cadastro_id}/aceites")
async def get_aceites_cliente(
    cadastro_id: str,
    usuario: dict = Depends(verificar_admin)
):
    """
    Lista todos os aceites de termos de um cliente.
    Usado para auditoria e comprovação legal.
    """
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão com banco de dados")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Buscar todos os aceites com informações da versão dos termos
        cur.execute("""
            SELECT 
                a.id,
                a.cadastro_id,
                t.tipo,
                t.versao,
                t.conteudo,
                a.ip_address,
                a.user_agent,
                a.aceito_em,
                a.metadados
            FROM aceites_termos a
            JOIN termos_versoes t ON a.termos_versao_id = t.id
            WHERE a.cadastro_id = %s
            ORDER BY a.aceito_em DESC
        """, (cadastro_id,))
        
        rows = cur.fetchall()
        cur.close()
        conn.close()
        
        aceites = []
        for row in rows:
            aceites.append({
                "id": row["id"],
                "cadastro_id": row["cadastro_id"],
                "tipo": row["tipo"],
                "versao": row["versao"],
                "conteudo_resumo": row["conteudo"][:200] + "..." if row["conteudo"] and len(row["conteudo"]) > 200 else row["conteudo"],
                "ip_address": row["ip_address"],
                "user_agent": row["user_agent"],
                "aceito_em": row["aceito_em"].isoformat() if row["aceito_em"] else None,
                "metadados": row["metadados"]
            })
        
        return {
            "cadastro_id": cadastro_id,
            "total_aceites": len(aceites),
            "aceites": aceites
        }
        
    except Exception as e:
        logger.error(f"Erro ao buscar aceites do cliente {cadastro_id}: {e}")
        raise HTTPException(status_code=500, detail="Erro ao buscar aceites")


@app.get("/api/admin/aceites/exportar/{cadastro_id}")
async def exportar_aceites_cliente(
    cadastro_id: str,
    usuario: dict = Depends(verificar_admin)
):
    """
    Exporta todos os aceites de um cliente em formato JSON completo.
    Útil para comprovação judicial.
    """
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão com banco de dados")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Buscar dados do cadastro
        cur.execute("SELECT * FROM cadastros WHERE id = %s", (cadastro_id,))
        cadastro = cur.fetchone()
        
        if not cadastro:
            raise HTTPException(status_code=404, detail="Cadastro não encontrado")
        
        # Buscar todos os aceites com o conteúdo completo dos termos
        cur.execute("""
            SELECT 
                a.id,
                a.cadastro_id,
                t.tipo,
                t.versao,
                t.conteudo,
                t.data_vigencia,
                a.ip_address,
                a.user_agent,
                a.aceito_em,
                a.metadados
            FROM aceites_termos a
            JOIN termos_versoes t ON a.termos_versao_id = t.id
            WHERE a.cadastro_id = %s
            ORDER BY a.aceito_em DESC
        """, (cadastro_id,))
        
        aceites = cur.fetchall()
        cur.close()
        conn.close()
        
        # Montar documento de comprovação
        dados_cadastro = cadastro["dados"] if isinstance(cadastro["dados"], dict) else json.loads(cadastro["dados"] or "{}")
        
        documento_comprovacao = {
            "titulo": "COMPROVAÇÃO DE ACEITE DE TERMOS",
            "gerado_em": datetime.now().isoformat(),
            "gerado_por": usuario["email"],
            "cliente": {
                "cadastro_id": cadastro_id,
                "nome": dados_cadastro.get("nome"),
                "cpf": dados_cadastro.get("cpf"),
                "email": dados_cadastro.get("email"),
                "data_cadastro": str(cadastro["data"]),
                "termos_aceitos": cadastro.get("termos_aceitos", False),
                "termos_aceitos_em": cadastro.get("termos_aceitos_em").isoformat() if cadastro.get("termos_aceitos_em") else None
            },
            "aceites": [
                {
                    "id_aceite": a["id"],
                    "tipo_documento": a["tipo"],
                    "versao_documento": a["versao"],
                    "data_vigencia_documento": a["data_vigencia"].isoformat() if a["data_vigencia"] else None,
                    "conteudo_integral": a["conteudo"],
                    "ip_address": a["ip_address"],
                    "user_agent": a["user_agent"],
                    "data_hora_aceite": a["aceito_em"].isoformat() if a["aceito_em"] else None,
                    "metadados_tecnicos": a["metadados"]
                }
                for a in aceites
            ],
            "declaracao": "Este documento comprova que o titular acima identificado aceitou os termos nas datas e condições especificadas, através do sistema digital do escritório Vaucher e Álvares Sociedade de Advogados."
        }
        
        return documento_comprovacao
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao exportar aceites: {e}")
        raise HTTPException(status_code=500, detail="Erro ao exportar aceites")

# ============================================
# ENDPOINTS - DEMANDA ESPECÍFICA E DOCUMENTOS
# ============================================

@app.post("/api/admin/clientes/{cadastro_id}/importar-documentos-demanda")
async def importar_documentos_para_demanda(
    cadastro_id: str,
    usuario: dict = Depends(verificar_admin)
):
    """Importa documentos existentes da pasta do cliente para a tabela documentos_demanda.
    
    Útil para migrar documentos de clientes que foram cadastrados antes da 
    funcionalidade de documentos específicos da demanda.
    """
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    # Verificar pastas de documentos
    pasta_cliente = os.path.join(UPLOADS_DIR, cadastro_id)
    pasta_demanda = os.path.join(UPLOADS_DIR, "documentos_demanda", cadastro_id)
    
    importados = []
    ja_existentes = []
    
    # Verificar documentos na pasta genérica do cliente
    if os.path.exists(pasta_cliente):
        for arquivo in os.listdir(pasta_cliente):
            arquivo_path = os.path.join(pasta_cliente, arquivo)
            if os.path.isfile(arquivo_path):
                # Determinar tipo baseado no nome do arquivo
                nome_lower = arquivo.lower()
                if 'contracheque' in nome_lower or 'holerite' in nome_lower or 'pagamento' in nome_lower:
                    tipo = 'contracheque'
                elif 'certificado' in nome_lower or 'residencia' in nome_lower or 'coreme' in nome_lower:
                    tipo = 'certificado_residencia'
                elif 'comprovante' in nome_lower or 'endereco' in nome_lower:
                    tipo = 'comprovante_residencia'
                elif 'rg' in nome_lower or 'cpf' in nome_lower or 'identidade' in nome_lower or 'cnh' in nome_lower:
                    tipo = 'documentos_pessoais'
                elif 'processo' in nome_lower or 'peticao' in nome_lower:
                    tipo = 'processo_anterior'
                else:
                    tipo = 'documento_geral'
                
                # Copiar para pasta de demanda
                os.makedirs(pasta_demanda, exist_ok=True)
                ext = os.path.splitext(arquivo)[1]
                novo_nome = f"{tipo}_{uuid.uuid4().hex[:8]}{ext}"
                novo_path = os.path.join(pasta_demanda, novo_nome)
                
                try:
                    shutil.copy2(arquivo_path, novo_path)
                    
                    # Salvar na tabela
                    sucesso = salvar_documento_demanda(
                        cadastro_id, tipo, novo_nome, 
                        arquivo, novo_path, f"Importado de {arquivo}"
                    )
                    
                    if sucesso:
                        importados.append({"nome": arquivo, "tipo": tipo})
                    else:
                        ja_existentes.append(arquivo)
                except Exception as e:
                    logger.error(f"Erro ao importar {arquivo}: {e}")
    
    # Também verificar lista de documentos do cadastro que podem não estar na pasta
    for doc_nome in cadastro.get("documentos", []):
        doc_path = os.path.join(pasta_cliente, doc_nome)
        if os.path.exists(doc_path) and doc_nome not in [d["nome"] for d in importados]:
            # Já foi processado no loop acima
            continue
    
    return {
        "success": True,
        "importados": importados,
        "total_importados": len(importados),
        "ja_existentes": ja_existentes,
        "message": f"{len(importados)} documento(s) importado(s) para a demanda"
    }


@app.post("/api/cadastros/{cadastro_id}/demanda-especifica")
async def salvar_rascunho_demanda(cadastro_id: str, dados: SalvarRascunhoDemanda):
    """Salva ou atualiza dados específicos de uma demanda (rascunho)."""
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    sucesso = salvar_dados_demanda(cadastro_id, dados.tipo_demanda, dados.dados, "rascunho")
    if not sucesso:
        raise HTTPException(status_code=500, detail="Erro ao salvar rascunho")
    
    return {"success": True, "message": "Rascunho salvo com sucesso"}

@app.get("/api/cadastros/{cadastro_id}/demanda-especifica/{tipo_demanda}")
async def buscar_rascunho_demanda(cadastro_id: str, tipo_demanda: str):
    """Busca dados específicos de uma demanda."""
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    dados = buscar_dados_demanda(cadastro_id, tipo_demanda)
    return {"success": True, "dados": dados}

@app.post("/api/cadastros/{cadastro_id}/documento-demanda/{tipo_documento}")
async def upload_documento_demanda(
    cadastro_id: str, 
    tipo_documento: str,
    arquivo: UploadFile = File(...),
    descricao: str = Form("")
):
    """Faz upload de um documento específico da demanda."""
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    # Criar diretório para documentos da demanda
    cliente_dir = os.path.join(UPLOADS_DIR, "documentos_demanda", cadastro_id)
    os.makedirs(cliente_dir, exist_ok=True)
    
    # Gerar nome único para o arquivo
    ext = os.path.splitext(arquivo.filename)[1]
    nome_arquivo = f"{tipo_documento}_{uuid.uuid4().hex[:8]}{ext}"
    arquivo_path = os.path.join(cliente_dir, nome_arquivo)
    
    # Salvar arquivo
    with open(arquivo_path, "wb") as f:
        content = await arquivo.read()
        f.write(content)
    
    # Salvar referência no banco
    sucesso = salvar_documento_demanda(
        cadastro_id, tipo_documento, nome_arquivo, 
        arquivo.filename, arquivo_path, descricao
    )
    
    if not sucesso:
        raise HTTPException(status_code=500, detail="Erro ao registrar documento")
    
    return {"success": True, "message": "Documento enviado com sucesso", "nome_arquivo": nome_arquivo}

@app.get("/api/cadastros/{cadastro_id}/documentos-demanda")
async def listar_docs_demanda(cadastro_id: str):
    """Lista documentos específicos da demanda."""
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    documentos = listar_documentos_demanda(cadastro_id)
    return {"success": True, "documentos": documentos}

@app.get("/api/cadastros/{cadastro_id}/documentos-demanda/{doc_id}/download")
async def download_documento_demanda(cadastro_id: str, doc_id: int):
    """Download de um documento específico da demanda."""
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Erro de conexão com banco")
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT * FROM documentos_demanda 
            WHERE id = %s AND cadastro_id = %s
        """, (doc_id, cadastro_id))
        doc = cur.fetchone()
        cur.close()
        conn.close()
        
        if not doc:
            raise HTTPException(status_code=404, detail="Documento não encontrado")
        
        arquivo_path = doc['arquivo_path']
        if not os.path.exists(arquivo_path):
            raise HTTPException(status_code=404, detail="Arquivo não encontrado no servidor")
        
        return FileResponse(
            arquivo_path,
            filename=doc['nome_original'],
            media_type="application/octet-stream"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao baixar documento da demanda: {e}")
        raise HTTPException(status_code=500, detail="Erro ao baixar documento")


# A função gerar_peticao_auxilio_moradia foi movida para modules/documents.py

# ============================================
# ENDPOINTS PARA GERAR PETIÇÃO INICIAL
# ============================================

@app.post("/api/admin/clientes/{cadastro_id}/gerar-peticao/{tipo_demanda}")
async def gerar_peticao_inicial(cadastro_id: str, tipo_demanda: str, usuario: dict = Depends(verificar_admin)):
    """Gera a petição inicial para uma demanda específica."""
    
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    dados_demanda = buscar_dados_demanda(cadastro_id, tipo_demanda)
    if not dados_demanda:
        raise HTTPException(status_code=404, detail="Dados da demanda não encontrados")
    
    try:
        if tipo_demanda == "auxilio_moradia_residencia":
            caminho = gerar_peticao_auxilio_moradia(
                cadastro.get('dados', {}),
                dados_demanda.get('dados', {}),
                cadastro_id
            )
            
            # Atualizar arquivos_gerados no cadastro
            arquivos = cadastro.get('arquivos_gerados', {})
            arquivos['peticao_auxilio_moradia'] = caminho
            
            conn = get_db()
            if conn:
                cur = conn.cursor()
                cur.execute("""
                    UPDATE cadastros SET arquivos_gerados = %s WHERE id = %s
                """, (json.dumps(arquivos), cadastro_id))
                conn.commit()
                cur.close()
                conn.close()
            
            return {
                "success": True,
                "message": "Petição gerada com sucesso",
                "arquivo": os.path.basename(caminho)
            }
        else:
            raise HTTPException(status_code=400, detail="Tipo de demanda não suportado para geração de petição")
    
    except Exception as e:
        logger.error(f"Erro ao gerar petição: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao gerar petição: {str(e)}")

@app.get("/api/admin/clientes/{cadastro_id}/peticao/{tipo_demanda}")
async def download_peticao(cadastro_id: str, tipo_demanda: str, usuario: dict = Depends(verificar_admin)):
    """Download da petição gerada."""
    
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")
    
    arquivos = cadastro.get('arquivos_gerados', {})
    
    if tipo_demanda == "auxilio_moradia_residencia":
        caminho = arquivos.get('peticao_auxilio_moradia')
    else:
        raise HTTPException(status_code=400, detail="Tipo de demanda não suportado")
    
    if not caminho or not os.path.exists(caminho):
        raise HTTPException(status_code=404, detail="Petição não encontrada. Gere a petição primeiro.")
    
    return FileResponse(
        caminho,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        filename=os.path.basename(caminho)
    )

# ============================================
# ASSINATURA DIGITAL - GOV.BR E ZAPSIGN
# ============================================

@app.get("/api/assinatura/govbr")
async def obter_link_govbr():
    """Retorna informações para assinatura via Gov.br."""
    return gerar_link_govbr()


@app.post("/api/admin/clientes/{cadastro_id}/upload-documento-final/{tipo_documento}")
async def upload_documento_final(
    cadastro_id: str,
    tipo_documento: str,
    arquivo: UploadFile = File(...),
    usuario: dict = Depends(verificar_admin)
):
    """
    Faz upload do documento final (editado) para assinatura.
    O admin deve baixar o documento gerado, editar (preencher honorários, etc),
    e fazer upload da versão final aqui antes de enviar para assinatura.

    tipo_documento: contrato, procuracao, prestacao_contas
    """
    if tipo_documento not in ["contrato", "procuracao", "prestacao_contas"]:
        raise HTTPException(status_code=400, detail="Tipo de documento deve ser 'contrato', 'procuracao' ou 'prestacao_contas'")

    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")

    # Verificar extensão do arquivo
    extensao = os.path.splitext(arquivo.filename)[1].lower()
    if extensao not in ['.pdf', '.docx', '.doc']:
        raise HTTPException(status_code=400, detail="Arquivo deve ser PDF ou DOCX")

    # Criar pasta para documentos finais (usando UPLOADS_DIR para consistência)
    pasta_finais = os.path.join(UPLOADS_DIR, "documentos_finais", cadastro_id)
    os.makedirs(pasta_finais, exist_ok=True)

    # Nome do arquivo final
    nome_arquivo = f"{tipo_documento}_final{extensao}"
    caminho_final = os.path.join(pasta_finais, nome_arquivo)

    # Salvar arquivo
    try:
        conteudo = await arquivo.read()
        with open(caminho_final, "wb") as f:
            f.write(conteudo)
        logger.info(f"Documento final salvo: {caminho_final}")
    except Exception as e:
        logger.error(f"Erro ao salvar documento final: {e}")
        raise HTTPException(status_code=500, detail="Erro ao salvar arquivo")

    # Atualizar cadastro com caminho do documento final
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()

            # Buscar documentos finais existentes ou criar novo dict
            documentos_finais = cadastro.get('documentos_finais', {})
            if not isinstance(documentos_finais, dict):
                documentos_finais = {}

            documentos_finais[tipo_documento] = {
                "caminho": caminho_final,
                "nome_original": arquivo.filename,
                "upload_em": datetime.now().isoformat(),
                "upload_por": usuario.get("email", "admin")
            }

            cur.execute("""
                UPDATE cadastros
                SET documentos_finais = %s,
                    atualizado_em = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (json.dumps(documentos_finais), cadastro_id))
            conn.commit()
            cur.close()
            conn.close()

            return {
                "success": True,
                "message": f"Documento {tipo_documento} final enviado com sucesso!",
                "arquivo": nome_arquivo
            }
        except Exception as e:
            logger.error(f"Erro ao atualizar cadastro: {e}")
            raise HTTPException(status_code=500, detail="Erro ao atualizar cadastro")

    raise HTTPException(status_code=500, detail="Erro de conexão com banco")


@app.get("/api/admin/clientes/{cadastro_id}/documentos-finais")
async def obter_documentos_finais(
    cadastro_id: str,
    usuario: dict = Depends(verificar_admin)
):
    """Retorna status dos documentos finais (editados) do cliente."""
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")

    documentos_finais = cadastro.get('documentos_finais', {})
    if not isinstance(documentos_finais, dict):
        documentos_finais = {}

    return {
        "success": True,
        "documentos_finais": documentos_finais,
        "tem_contrato_final": "contrato" in documentos_finais,
        "tem_procuracao_final": "procuracao" in documentos_finais
    }


@app.get("/api/admin/clientes/{cadastro_id}/download-documento-final/{tipo_documento}")
async def download_documento_final(
    cadastro_id: str,
    tipo_documento: str,
    usuario: dict = Depends(verificar_admin)
):
    """Baixa o documento final (editado)."""
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")

    documentos_finais = cadastro.get('documentos_finais', {})
    if tipo_documento not in documentos_finais:
        raise HTTPException(status_code=404, detail="Documento final não encontrado")

    caminho = documentos_finais[tipo_documento].get("caminho")
    if not caminho or not os.path.exists(caminho):
        raise HTTPException(status_code=404, detail="Arquivo não encontrado")

    nome_original = documentos_finais[tipo_documento].get("nome_original", f"{tipo_documento}_final.pdf")

    return FileResponse(
        caminho,
        filename=nome_original,
        media_type="application/octet-stream"
    )


@app.post("/api/admin/clientes/{cadastro_id}/enviar-assinatura/{tipo_documento}")
async def enviar_documento_assinatura(
    cadastro_id: str,
    tipo_documento: str,
    usuario: dict = Depends(verificar_admin)
):
    """
    Envia documento para assinatura digital via ZapSign.

    IMPORTANTE: Para contrato e procuração, é necessário fazer upload do documento
    final (editado com honorários preenchidos) antes de enviar para assinatura.

    tipo_documento: contrato, procuracao, peticao_auxilio_moradia
    """
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")

    dados = cadastro.get('dados', {})
    arquivos_gerados = cadastro.get('arquivos_gerados', {})
    documentos_finais = cadastro.get('documentos_finais', {})
    if not isinstance(documentos_finais, dict):
        documentos_finais = {}

    # Determinar qual arquivo enviar
    # PRIORIDADE: documento final (editado) > documento gerado automaticamente
    caminho = None
    nome_doc = ""

    if tipo_documento == "contrato":
        # Para contrato, OBRIGATÓRIO ter documento final (com honorários preenchidos)
        if "contrato" in documentos_finais:
            caminho = documentos_finais["contrato"].get("caminho")
        if not caminho or not os.path.exists(caminho):
            raise HTTPException(
                status_code=400,
                detail="É necessário fazer upload do contrato editado (com honorários preenchidos) antes de enviar para assinatura. Baixe o contrato, edite e faça upload da versão final."
            )
        nome_doc = f"Contrato de Honorários - {dados.get('nome', 'Cliente')}"

    elif tipo_documento == "procuracao":
        # Para procuração, OBRIGATÓRIO ter documento final
        if "procuracao" in documentos_finais:
            caminho = documentos_finais["procuracao"].get("caminho")
        if not caminho or not os.path.exists(caminho):
            raise HTTPException(
                status_code=400,
                detail="É necessário fazer upload da procuração editada antes de enviar para assinatura. Baixe a procuração, revise e faça upload da versão final."
            )
        nome_doc = f"Procuração - {dados.get('nome', 'Cliente')}"

    elif tipo_documento == "peticao_auxilio_moradia":
        # Para petição, pode usar o gerado diretamente
        caminho = arquivos_gerados.get('peticao_auxilio_moradia')
        nome_doc = f"Petição Auxílio Moradia - {dados.get('nome', 'Cliente')}"
        if not caminho or not os.path.exists(caminho):
            raise HTTPException(
                status_code=404,
                detail="Petição não encontrada. Gere a petição primeiro."
            )

    elif tipo_documento == "prestacao_contas":
        # Para prestação de contas, OBRIGATÓRIO ter documento final (revisado)
        if "prestacao_contas" in documentos_finais:
            caminho = documentos_finais["prestacao_contas"].get("caminho")
        if not caminho or not os.path.exists(caminho):
            raise HTTPException(
                status_code=400,
                detail="É necessário fazer upload da prestação de contas revisada antes de enviar para assinatura. Gere o documento, revise e faça upload da versão final."
            )
        nome_doc = f"Prestação de Contas - {dados.get('nome', 'Cliente')}"

    else:
        raise HTTPException(status_code=400, detail="Tipo de documento não suportado")

    # Verificar se o arquivo existe
    logger.info(f"=== ENVIO PARA ZAPSIGN ===")
    logger.info(f"Tipo documento: {tipo_documento}")
    logger.info(f"Caminho arquivo: {caminho}")
    logger.info(f"Arquivo existe: {os.path.exists(caminho) if caminho else 'N/A'}")
    if caminho and os.path.exists(caminho):
        file_size = os.path.getsize(caminho)
        logger.info(f"Tamanho arquivo: {file_size} bytes")
    else:
        logger.error(f"ARQUIVO NÃO ENCONTRADO: {caminho}")

    # Preparar signatário (cliente)
    signatarios = [{
        "nome": dados.get('nome', ''),
        "email": dados.get('email', ''),
        "telefone": dados.get('telefone', ''),
        "auth_mode": "assinaturaTela"
    }]

    logger.info(f"Signatário: {dados.get('nome')} - {dados.get('email')}")

    # Enviar para ZapSign
    resultado = await criar_documento_zapsign(
        nome_documento=nome_doc,
        arquivo_path=caminho,
        signatarios=signatarios,
        enviar_email_automatico=True
    )

    if not resultado.get("success"):
        error_msg = resultado.get("error", "Erro desconhecido ao enviar para ZapSign")
        logger.error(f"FALHA ao enviar para ZapSign: {error_msg}")
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao enviar para ZapSign: {error_msg}"
        )

    url_assinatura = resultado.get("signatarios", [{}])[0].get("url_assinatura")

    logger.info(f"ZapSign resultado: {resultado}")
    logger.info(f"URL de assinatura obtida: {url_assinatura}")

    if not url_assinatura:
        logger.warning("URL de assinatura não retornada pelo ZapSign!")

    # Salvar token do documento no cadastro
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()

            # Buscar assinaturas existentes ou criar novo dict
            assinaturas = cadastro.get('assinaturas_digitais', {})
            assinaturas[tipo_documento] = {
                "token": resultado.get("documento_token"),
                "status": "pending",
                "url_assinatura": url_assinatura,
                "enviado_em": datetime.now().isoformat(),
                "plataforma": "zapsign"
            }

            cur.execute("""
                UPDATE cadastros
                SET assinaturas_digitais = %s,
                    atualizado_em = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (json.dumps(assinaturas), cadastro_id))
            conn.commit()
            cur.close()
            conn.close()
        except Exception as e:
            logger.error(f"Erro ao salvar assinatura: {e}")

    # Enviar e-mail para o cliente com link de assinatura E anexo
    email_enviado = False
    try:
        documentos_email = [{
            "tipo": tipo_documento,
            "nome": nome_doc,
            "url_assinatura": url_assinatura
        }]

        logger.info(f"Enviando e-mail com documentos: {documentos_email}")

        # Preparar anexo do documento
        anexos_email = []
        if caminho and os.path.exists(caminho):
            try:
                with open(caminho, "rb") as f:
                    conteudo_base64 = base64.b64encode(f.read()).decode("utf-8")

                # Determinar nome do arquivo
                extensao = os.path.splitext(caminho)[1]
                nome_anexo = f"{tipo_documento}{extensao}"

                anexos_email.append({
                    "filename": nome_anexo,
                    "content": conteudo_base64
                })
                logger.info(f"Anexo preparado: {nome_anexo}")
            except Exception as e:
                logger.error(f"Erro ao preparar anexo: {e}")

        email_enviado = await enviar_email_assinatura_digital(
            destinatario=dados.get('email', ''),
            nome=dados.get('nome', 'Cliente'),
            documentos=documentos_email,
            anexos=anexos_email if anexos_email else None
        )

        if email_enviado:
            logger.info(f"E-mail de assinatura enviado para {dados.get('email')}")
        else:
            logger.warning(f"Falha ao enviar e-mail de assinatura para {dados.get('email')}")
    except Exception as e:
        logger.error(f"Erro ao enviar e-mail de assinatura: {e}")

    return {
        "success": True,
        "message": "Documento enviado para assinatura e e-mail enviado ao cliente",
        "url_assinatura": url_assinatura,
        "documento_token": resultado.get("documento_token"),
        "email_enviado": email_enviado
    }


@app.get("/api/admin/clientes/{cadastro_id}/status-assinatura/{tipo_documento}")
async def verificar_status_assinatura(
    cadastro_id: str,
    tipo_documento: str,
    usuario: dict = Depends(verificar_admin)
):
    """Verifica o status de assinatura de um documento."""
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")

    assinaturas = cadastro.get('assinaturas_digitais', {})
    assinatura = assinaturas.get(tipo_documento)

    if not assinatura:
        return {"status": "nao_enviado", "message": "Documento não foi enviado para assinatura"}

    if assinatura.get("plataforma") == "zapsign":
        token = assinatura.get("token")
        if token:
            resultado = await verificar_status_documento(token)
            if resultado.get("success"):
                # Atualizar status no banco se mudou
                if resultado.get("todos_assinaram") and assinatura.get("status") != "signed":
                    conn = get_db()
                    if conn:
                        try:
                            cur = conn.cursor()
                            assinaturas[tipo_documento]["status"] = "signed"
                            cur.execute("""
                                UPDATE cadastros
                                SET assinaturas_digitais = %s
                                WHERE id = %s
                            """, (json.dumps(assinaturas), cadastro_id))
                            conn.commit()
                            cur.close()
                            conn.close()
                        except Exception as e:
                            logger.error(f"Erro ao atualizar status: {e}")

                return resultado

    return assinatura


@app.get("/api/admin/clientes/{cadastro_id}/documento-assinado/{tipo_documento}")
async def download_documento_assinado(
    cadastro_id: str,
    tipo_documento: str,
    usuario: dict = Depends(verificar_admin)
):
    """Obtém URL do documento assinado."""
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")

    assinaturas = cadastro.get('assinaturas_digitais', {})
    assinatura = assinaturas.get(tipo_documento)

    if not assinatura:
        raise HTTPException(status_code=404, detail="Documento não foi enviado para assinatura")

    if assinatura.get("plataforma") == "zapsign":
        token = assinatura.get("token")
        if token:
            resultado = await obter_documento_assinado(token)
            if resultado.get("success"):
                return resultado
            else:
                raise HTTPException(status_code=400, detail=resultado.get("error"))

    raise HTTPException(status_code=400, detail="Não foi possível obter documento assinado")


@app.post("/api/admin/clientes/{cadastro_id}/verificar-e-baixar-assinatura/{tipo_documento}")
async def verificar_e_baixar_assinatura(
    cadastro_id: str,
    tipo_documento: str,
    usuario: dict = Depends(verificar_admin)
):
    """
    Verifica status da assinatura no ZapSign.
    Se assinado, baixa o documento e salva localmente.
    """
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")

    assinaturas = cadastro.get('assinaturas_digitais', {})
    assinatura = assinaturas.get(tipo_documento)

    if not assinatura:
        return {
            "success": False,
            "status": "nao_enviado",
            "message": "Documento nao foi enviado para assinatura"
        }

    if assinatura.get("plataforma") != "zapsign":
        return {
            "success": False,
            "status": "erro",
            "message": "Plataforma nao suportada"
        }

    token = assinatura.get("token")
    if not token:
        return {
            "success": False,
            "status": "erro",
            "message": "Token nao encontrado"
        }

    # Verificar status no ZapSign
    resultado = await verificar_status_documento(token)

    if not resultado.get("success"):
        return {
            "success": False,
            "status": "erro",
            "message": resultado.get("error", "Erro ao verificar status")
        }

    status_atual = resultado.get("status")
    todos_assinaram = resultado.get("todos_assinaram", False)

    # Se ainda nao assinado
    if not todos_assinaram:
        return {
            "success": True,
            "status": "pending",
            "message": "Documento ainda nao foi assinado",
            "detalhes": resultado.get("signatarios", [])
        }

    # Documento foi assinado! Vamos baixar
    logger.info(f"Documento {tipo_documento} do cadastro {cadastro_id} foi assinado! Baixando...")

    # Obter URL do documento assinado
    doc_assinado = await obter_documento_assinado(token)

    if not doc_assinado.get("success"):
        return {
            "success": False,
            "status": "signed",
            "message": "Documento assinado mas erro ao obter URL: " + doc_assinado.get("error", "")
        }

    url_documento = doc_assinado.get("url_documento_assinado")

    if not url_documento:
        return {
            "success": False,
            "status": "signed",
            "message": "URL do documento assinado nao disponivel"
        }

    # Baixar o documento
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(url_documento, timeout=60.0)

            if response.status_code != 200:
                return {
                    "success": False,
                    "status": "signed",
                    "message": f"Erro ao baixar documento: HTTP {response.status_code}"
                }

            # Criar pasta para documentos assinados (consistente com download)
            pasta_assinados = os.path.join(UPLOADS_DIR, "documentos_assinados", cadastro_id)
            os.makedirs(pasta_assinados, exist_ok=True)

            # Nome do arquivo
            nome_arquivo = f"{tipo_documento}_assinado.pdf"
            caminho_completo = os.path.join(pasta_assinados, nome_arquivo)
            logger.info(f"Salvando documento assinado em: {caminho_completo}")

            # Salvar arquivo
            with open(caminho_completo, "wb") as f:
                f.write(response.content)

            logger.info(f"Documento assinado salvo: {caminho_completo}")

            # Atualizar banco de dados
            conn = get_db()
            if conn:
                try:
                    cur = conn.cursor()

                    # Atualizar assinaturas_digitais
                    assinaturas[tipo_documento]["status"] = "signed"
                    assinaturas[tipo_documento]["arquivo_assinado"] = caminho_completo
                    assinaturas[tipo_documento]["assinado_em"] = datetime.now().isoformat()

                    # Atualizar documentos_assinados (lista)
                    docs_assinados = cadastro.get("documentos_assinados", [])
                    if not isinstance(docs_assinados, list):
                        docs_assinados = []
                    if nome_arquivo not in docs_assinados:
                        docs_assinados.append(nome_arquivo)

                    cur.execute("""
                        UPDATE cadastros
                        SET assinaturas_digitais = %s,
                            documentos_assinados = %s,
                            data_assinatura = CURRENT_TIMESTAMP,
                            atualizado_em = CURRENT_TIMESTAMP
                        WHERE id = %s
                    """, (json.dumps(assinaturas), json.dumps(docs_assinados), cadastro_id))
                    conn.commit()
                    cur.close()
                    conn.close()

                    logger.info(f"Banco atualizado para cadastro {cadastro_id}")

                except Exception as e:
                    logger.error(f"Erro ao atualizar banco: {e}")

            return {
                "success": True,
                "status": "signed",
                "message": "Documento assinado baixado com sucesso!",
                "arquivo": nome_arquivo,
                "caminho": caminho_completo
            }

    except Exception as e:
        logger.error(f"Erro ao baixar documento assinado: {e}")
        return {
            "success": False,
            "status": "signed",
            "message": f"Erro ao baixar: {str(e)}"
        }


@app.post("/api/admin/clientes/{cadastro_id}/enviar-email-assinatura")
async def enviar_email_com_assinatura(
    cadastro_id: str,
    usuario: dict = Depends(verificar_admin)
):
    """Envia e-mail com links para assinatura digital (ZapSign + Gov.br)."""
    cadastro = buscar_cadastro(cadastro_id)
    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")

    dados = cadastro.get('dados', {})
    email = dados.get('email')
    nome = dados.get('nome', 'Cliente')

    if not email:
        raise HTTPException(status_code=400, detail="Cliente não possui e-mail cadastrado")

    assinaturas = cadastro.get('assinaturas_digitais', {})

    # Montar corpo do e-mail
    corpo_html = f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
        <div style="background-color: #8B1538; padding: 20px; text-align: center;">
            <h1 style="color: white; margin: 0;">Vaucher e Álvares</h1>
            <p style="color: white; margin: 5px 0;">Sociedade de Advogados</p>
        </div>

        <div style="padding: 30px; background-color: #f9f9f9;">
            <h2 style="color: #333;">Olá, {nome}!</h2>

            <p style="color: #555; line-height: 1.6;">
                Seus documentos estão prontos para assinatura digital.
                Você pode assinar de duas formas:
            </p>
    """

    # Adicionar links de assinatura ZapSign se existirem
    docs_pendentes = []
    for tipo, info in assinaturas.items():
        if info.get("status") == "pending" and info.get("url_assinatura"):
            nome_doc = tipo.replace("_", " ").title()
            docs_pendentes.append({
                "nome": nome_doc,
                "url": info.get("url_assinatura")
            })

    if docs_pendentes:
        corpo_html += """
            <h3 style="color: #8B1538; margin-top: 30px;">Opção 1: Assinatura Rápida (ZapSign)</h3>
            <p style="color: #555;">Clique nos links abaixo para assinar cada documento:</p>
        """
        for doc in docs_pendentes:
            corpo_html += f"""
            <p style="margin: 10px 0;">
                <a href="{doc['url']}"
                   style="display: inline-block;
                          background-color: #8B1538;
                          color: white;
                          padding: 12px 25px;
                          text-decoration: none;
                          border-radius: 5px;
                          font-weight: bold;">
                    Assinar {doc['nome']}
                </a>
            </p>
            """

    # Adicionar opção Gov.br
    corpo_html += f"""
            <h3 style="color: #1351B4; margin-top: 30px;">Opção 2: Assinatura via Gov.br</h3>
            <p style="color: #555;">
                Se preferir usar sua conta Gov.br (nível Prata ou Ouro):
            </p>
            <ol style="color: #555; line-height: 1.8;">
                <li>Baixe os documentos no Portal do Cliente</li>
                <li>Acesse o Assinador Gov.br clicando no botão abaixo</li>
                <li>Faça login com sua conta Gov.br</li>
                <li>Faça upload e assine cada documento</li>
                <li>Envie os documentos assinados pelo Portal do Cliente</li>
            </ol>

            <p style="margin: 20px 0;">
                <a href="https://sso.acesso.gov.br/login?client_id=assinador.iti.br"
                   style="display: inline-block;
                          background-color: #1351B4;
                          color: white;
                          padding: 12px 25px;
                          text-decoration: none;
                          border-radius: 5px;
                          font-weight: bold;">
                    Acessar Assinador Gov.br
                </a>
            </p>
        </div>

        <div style="background-color: #333; padding: 20px; text-align: center;">
            <p style="color: #999; font-size: 12px; margin: 0;">
                Vaucher e Álvares Sociedade de Advogados<br>
                OAB/MT 669 | CNPJ 21.336.697/0001-46<br>
                Rua Lima, n. 106, Jardim das Américas - Cuiabá/MT
            </p>
        </div>
    </div>
    """

    # Enviar e-mail
    sucesso = await enviar_email_resend(
        destinatario=email,
        assunto="Documentos para Assinatura Digital - Vaucher e Álvares",
        corpo_html=corpo_html
    )

    if sucesso:
        return {"success": True, "message": f"E-mail enviado para {email}"}
    else:
        raise HTTPException(status_code=500, detail="Erro ao enviar e-mail")


# Endpoint para cliente obter seus links de assinatura
@app.get("/api/cliente/minhas-assinaturas")
async def cliente_minhas_assinaturas(cliente: dict = Depends(verificar_token_cliente)):
    """Retorna links de assinatura pendentes do cliente."""
    cadastro_id = cliente.get("cadastro_id")
    cadastro = buscar_cadastro(cadastro_id)

    if not cadastro:
        raise HTTPException(status_code=404, detail="Cadastro não encontrado")

    assinaturas = cadastro.get('assinaturas_digitais', {})
    govbr = gerar_link_govbr()

    logger.info(f"Buscando assinaturas para cadastro {cadastro_id}: {assinaturas}")

    # Lista de assinaturas pendentes para o frontend
    lista_assinaturas = []
    for tipo, info in assinaturas.items():
        # Incluir assinaturas pendentes que tenham URL
        if isinstance(info, dict) and info.get("status") == "pending":
            url = info.get("url_assinatura")
            if url:
                lista_assinaturas.append({
                    "tipo": tipo,
                    "url": url,
                    "status": info.get("status")
                })
                logger.info(f"Assinatura pendente encontrada: {tipo} - {url}")

    logger.info(f"Total de assinaturas pendentes: {len(lista_assinaturas)}")

    return {
        "success": True,
        "assinaturas": lista_assinaturas,
        "govbr": govbr
    }


# ============================================
# IMPORTAÇÃO DO ASTREA
# ============================================

from openpyxl import load_workbook

def parsear_excel_astrea(arquivo_bytes: bytes) -> dict:
    """
    Parseia arquivo Excel do Astrea e retorna dados estruturados.
    Tenta detectar automaticamente as colunas baseado nos headers.
    """
    from io import BytesIO

    try:
        wb = load_workbook(filename=BytesIO(arquivo_bytes), read_only=True, data_only=True)
        ws = wb.active

        # Ler todas as linhas
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return {"erro": "Arquivo vazio ou sem dados"}

        # Primeira linha são os headers
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]

        # Mapeamento flexível de colunas do Astrea para campos do sistema
        # IMPORTANTE: Ordem importa - primeiro os mais específicos
        mapeamento_processos = {
            "numero_processo": ["número", "numero", "nº processo", "n° processo", "processo"],
            "tipo_acao": ["ação", "acao", "matéria", "materia", "tipo", "classe"],
            "vara_tribunal": ["vara", "tribunal", "órgão julgador", "orgao julgador"],
            "foro": ["foro", "comarca", "localidade"],
            "fase": ["instância atual", "instancia atual", "fase", "situação", "situacao", "status", "fase atual", "instância", "instancia"],
            "reu": ["outros envolvidos", "réu", "reu", "requerido", "reclamado", "polo passivo", "parte contrária", "parte contraria", "réus", "reus", "partes"],
            "valor_causa": ["valor da causa", "valor"],
            "data_distribuicao": ["data de distribuição", "data de distribuicao", "data distribuição", "data distribuicao", "distribuição", "distribuicao"],
            "observacoes": ["observações", "observacoes", "detalhes", "obs"],
            "cliente_nome": ["cliente", "nome do cliente", "parte", "autor", "requerente", "reclamante", "polo ativo", "partes", "nome"],
            "cliente_cpf": ["cpf", "cpf do cliente", "cpf cliente", "documento"],
            "titulo": ["título", "titulo", "assunto"],
            "pasta": ["pasta", "código", "codigo"],
            "objeto": ["objeto", "pedido"]
        }

        mapeamento_andamentos = {
            "data": ["data do último histórico", "data do ultimo historico", "data histórico", "data historico", "data andamento", "data", "data do andamento", "data movimentação", "data movimentacao"],
            "descricao": ["descrição do último histórico", "descricao do ultimo historico", "descrição histórico", "descricao historico", "descrição", "descricao", "histórico", "historico", "andamento", "movimentação", "movimentacao", "último andamento", "ultimo andamento"]
        }

        def encontrar_coluna(mapeamento_lista, headers):
            """Encontra o índice da coluna baseado no nome exato ou parcial."""
            # Primeiro: busca por match exato
            for idx, header in enumerate(headers):
                header_limpo = header.lower().strip()
                for possivel in mapeamento_lista:
                    if header_limpo == possivel:
                        return idx
            # Segundo: busca por match parcial (começa com)
            for idx, header in enumerate(headers):
                header_limpo = header.lower().strip()
                for possivel in mapeamento_lista:
                    if header_limpo.startswith(possivel) or possivel.startswith(header_limpo):
                        return idx
            return None

        # Detectar colunas de processos
        colunas_processo = {}
        for campo, variacoes in mapeamento_processos.items():
            idx = encontrar_coluna(variacoes, headers)
            if idx is not None:
                colunas_processo[campo] = idx

        # Detectar colunas de andamentos
        colunas_andamento = {}
        for campo, variacoes in mapeamento_andamentos.items():
            idx = encontrar_coluna(variacoes, headers)
            if idx is not None:
                colunas_andamento[campo] = idx

        # Logar colunas detectadas para debug
        logger.info(f"Headers Excel: {headers[:20]}")
        logger.info(f"Colunas de processo detectadas: {colunas_processo}")
        logger.info(f"Colunas de andamento detectadas: {colunas_andamento}")

        # Verificar especificamente coluna de cliente
        if "cliente_nome" in colunas_processo:
            logger.info(f"Coluna de cliente encontrada no índice {colunas_processo['cliente_nome']}: '{headers[colunas_processo['cliente_nome']]}'")
        else:
            logger.warning("Coluna de cliente NÃO encontrada no Excel!")
            logger.warning(f"Headers disponíveis: {headers}")

        # Verificar se encontrou coluna essencial (número do processo)
        if "numero_processo" not in colunas_processo:
            return {
                "erro": "Não foi possível identificar a coluna de número do processo",
                "colunas_detectadas": list(colunas_processo.keys()),
                "headers_encontrados": headers[:20]  # Primeiros 20 headers
            }

        # Processar dados
        processos_dict = {}  # Agrupa por numero_processo

        for row in rows[1:]:
            if not row or not any(row):
                continue

            def get_valor(campo, colunas, row):
                if campo in colunas:
                    idx = colunas[campo]
                    if idx < len(row):
                        val = row[idx]
                        return str(val).strip() if val else ""
                return ""

            numero_processo = get_valor("numero_processo", colunas_processo, row)
            if not numero_processo:
                continue

            # Se processo não existe no dict, criar
            if numero_processo not in processos_dict:
                # Converter valor_causa para decimal
                valor_causa_str = get_valor("valor_causa", colunas_processo, row)
                valor_causa = 0
                if valor_causa_str:
                    try:
                        # Remove R$, pontos de milhar, troca vírgula por ponto
                        valor_limpo = valor_causa_str.replace("R$", "").replace(".", "").replace(",", ".").strip()
                        valor_causa = float(valor_limpo)
                    except:
                        pass

                # Converter data_distribuicao
                data_dist = get_valor("data_distribuicao", colunas_processo, row)
                data_distribuicao = None
                if data_dist:
                    try:
                        # Tentar vários formatos de data
                        from datetime import datetime
                        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"]:
                            try:
                                data_distribuicao = datetime.strptime(data_dist, fmt).strftime("%Y-%m-%d")
                                break
                            except:
                                continue
                        # Se é um objeto datetime do Excel
                        if not data_distribuicao and isinstance(row[colunas_processo.get("data_distribuicao", -1)], datetime):
                            data_distribuicao = row[colunas_processo["data_distribuicao"]].strftime("%Y-%m-%d")
                    except:
                        pass

                    # Combinar Vara + Foro para vara_tribunal
                vara = get_valor("vara_tribunal", colunas_processo, row)
                foro = get_valor("foro", colunas_processo, row)
                vara_tribunal = vara
                if foro and foro != vara:
                    vara_tribunal = f"{vara} - {foro}" if vara else foro

                # Extrair réu dos "Outros envolvidos" - pegar quem é Requerido
                outros_envolvidos = get_valor("reu", colunas_processo, row)
                reu = ""
                if outros_envolvidos:
                    # Tentar extrair nomes com papel de Requerido
                    partes = outros_envolvidos.split("),")
                    reus = []
                    for parte in partes:
                        parte = parte.strip()
                        if "(Requerido" in parte or "(Réu" in parte:
                            # Extrair nome antes do parêntese
                            nome = parte.split("(")[0].strip()
                            if nome:
                                reus.append(nome)
                    reu = ", ".join(reus) if reus else outros_envolvidos

                # Pegar título se disponível (para observações)
                titulo = get_valor("titulo", colunas_processo, row)
                objeto = get_valor("objeto", colunas_processo, row)
                obs = get_valor("observacoes", colunas_processo, row)

                # Combinar observações
                obs_partes = []
                if objeto:
                    obs_partes.append(f"Objeto: {objeto}")
                if obs:
                    obs_partes.append(obs)
                observacoes_final = " | ".join(obs_partes) if obs_partes else ""

                processos_dict[numero_processo] = {
                    "numero_processo": numero_processo,
                    "tipo_acao": get_valor("tipo_acao", colunas_processo, row),
                    "vara_tribunal": vara_tribunal,
                    "fase": get_valor("fase", colunas_processo, row) or "Inicial",
                    "reu": reu,
                    "valor_causa": valor_causa,
                    "data_distribuicao": data_distribuicao,
                    "observacoes": observacoes_final,
                    "cliente_nome": get_valor("cliente_nome", colunas_processo, row),
                    "cliente_cpf": get_valor("cliente_cpf", colunas_processo, row),
                    "titulo": titulo,
                    "pasta": get_valor("pasta", colunas_processo, row),
                    "andamentos": []
                }
                # Log primeiros 3 processos para debug detalhado
                if len(processos_dict) <= 3:
                    logger.info(f"Processo #{len(processos_dict)}: numero={numero_processo}")
                    logger.info(f"  - cliente_nome='{get_valor('cliente_nome', colunas_processo, row)}'")
                    logger.info(f"  - fase='{get_valor('fase', colunas_processo, row)}'")
                    logger.info(f"  - reu='{reu}'")
                    logger.info(f"  - tipo_acao='{get_valor('tipo_acao', colunas_processo, row)}'")


            # Adicionar andamento se houver dados
            data_andamento = get_valor("data", colunas_andamento, row)
            descricao_andamento = get_valor("descricao", colunas_andamento, row)

            if data_andamento or descricao_andamento:
                # Converter data do andamento
                data_and_formatada = None
                if data_andamento:
                    try:
                        from datetime import datetime
                        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"]:
                            try:
                                data_and_formatada = datetime.strptime(data_andamento, fmt).strftime("%Y-%m-%d")
                                break
                            except:
                                continue
                        # Se é um objeto datetime do Excel
                        if not data_and_formatada and "data" in colunas_andamento:
                            idx = colunas_andamento["data"]
                            if idx < len(row) and isinstance(row[idx], datetime):
                                data_and_formatada = row[idx].strftime("%Y-%m-%d")
                    except:
                        pass

                andamento = {
                    "data": data_and_formatada or datetime.now().strftime("%Y-%m-%d"),
                    "descricao": descricao_andamento or "Sem descrição"
                }

                # Evitar duplicatas de andamentos
                existe = any(
                    a["data"] == andamento["data"] and a["descricao"] == andamento["descricao"]
                    for a in processos_dict[numero_processo]["andamentos"]
                )
                if not existe and descricao_andamento:
                    processos_dict[numero_processo]["andamentos"].append(andamento)

        processos_lista = list(processos_dict.values())
        total_andamentos = sum(len(p["andamentos"]) for p in processos_lista)

        wb.close()

        # Montar lista de colunas encontradas para exibição
        colunas_encontradas = list(colunas_processo.keys()) + list(colunas_andamento.keys())

        # Criar mapeamento de colunas encontradas com índices para debug
        colunas_mapeadas = {}
        for campo, idx in colunas_processo.items():
            colunas_mapeadas[campo] = {
                "indice": idx,
                "header_excel": headers[idx] if idx < len(headers) else "?"
            }
        for campo, idx in colunas_andamento.items():
            colunas_mapeadas[f"andamento_{campo}"] = {
                "indice": idx,
                "header_excel": headers[idx] if idx < len(headers) else "?"
            }

        return {
            "sucesso": True,
            "processos": processos_lista,
            "total_processos": len(processos_lista),
            "total_andamentos": total_andamentos,
            "colunas_encontradas": colunas_encontradas,
            "colunas_detectadas": {
                "processos": list(colunas_processo.keys()),
                "andamentos": list(colunas_andamento.keys())
            },
            "colunas_mapeadas": colunas_mapeadas,
            "headers_encontrados": headers[:30],
            "erros": []  # Erros serão adicionados durante o enriquecimento
        }

    except Exception as e:
        logger.error(f"Erro ao parsear Excel: {e}")
        return {"erro": f"Erro ao processar arquivo: {str(e)}"}


def buscar_processo_por_numero(numero_processo: str) -> dict:
    """Busca um processo pelo número."""
    conn = get_db()
    if not conn:
        return None

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM processos WHERE numero_processo = %s", (numero_processo,))
        row = cur.fetchone()
        cur.close()
        conn.close()

        if row:
            return dict(row)
        return None
    except Exception as e:
        logger.error(f"Erro ao buscar processo por número: {e}")
        return None


def buscar_cliente_por_cpf(cpf: str) -> dict:
    """Busca cliente pelo CPF."""
    conn = get_db()
    if not conn:
        return None

    try:
        # Limpar CPF - remover pontos e traços
        cpf_limpo = cpf.replace(".", "").replace("-", "").replace(" ", "")

        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, dados->>'nome' as nome, dados->>'cpf' as cpf
            FROM cadastros
            WHERE REPLACE(REPLACE(REPLACE(dados->>'cpf', '.', ''), '-', ''), ' ', '') = %s
        """, (cpf_limpo,))
        row = cur.fetchone()
        cur.close()
        conn.close()

        if row:
            return dict(row)
        return None
    except Exception as e:
        logger.error(f"Erro ao buscar cliente por CPF: {e}")
        return None


def normalizar_nome(nome: str) -> str:
    """Normaliza nome para comparação (uppercase, sem acentos)."""
    import unicodedata
    if not nome:
        return ""
    # Remove acentos
    nome_norm = unicodedata.normalize('NFKD', nome)
    nome_norm = ''.join(c for c in nome_norm if not unicodedata.combining(c))
    # Uppercase e remove espaços extras
    return ' '.join(nome_norm.upper().split())


def buscar_cliente_por_nome_fuzzy(nome: str, threshold: int = 80) -> dict:
    """
    Busca cliente usando fuzzy matching com rapidfuzz.
    threshold: mínimo de similaridade (0-100) para considerar match.
    """
    from rapidfuzz import fuzz, process

    conn = get_db()
    if not conn:
        return None

    try:
        nome_normalizado = normalizar_nome(nome)
        if not nome_normalizado or len(nome_normalizado) < 3:
            return None

        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, dados->>'nome' as nome, dados->>'cpf' as cpf
            FROM cadastros
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        if not rows:
            return None

        # Criar dicionário de nomes normalizados -> cliente
        clientes_map = {}
        for row in rows:
            nome_cliente = normalizar_nome(row.get("nome", ""))
            if nome_cliente:
                clientes_map[nome_cliente] = dict(row)

        if not clientes_map:
            return None

        # Usar rapidfuzz para encontrar melhor match
        resultado = process.extractOne(
            nome_normalizado,
            clientes_map.keys(),
            scorer=fuzz.token_sort_ratio  # Ignora ordem das palavras
        )

        if resultado:
            nome_match, score, _ = resultado
            logger.info(f"Fuzzy match: '{nome}' -> '{nome_match}' (score: {score})")
            if score >= threshold:
                cliente = clientes_map[nome_match]
                cliente["match_score"] = score
                return cliente

        return None
    except Exception as e:
        logger.error(f"Erro ao buscar cliente por nome (fuzzy): {e}")
        return None


def buscar_cliente_por_nome(nome: str) -> dict:
    """Busca cliente pelo nome usando fuzzy matching com threshold alto (quase exato)."""
    return buscar_cliente_por_nome_fuzzy(nome, threshold=95)


def buscar_clientes_similares(nome: str, limite: int = 5) -> list:
    """Busca clientes com nomes similares usando rapidfuzz."""
    from rapidfuzz import fuzz, process

    conn = get_db()
    if not conn:
        return []

    try:
        nome_normalizado = normalizar_nome(nome)
        if not nome_normalizado:
            return []

        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, dados->>'nome' as nome, dados->>'cpf' as cpf
            FROM cadastros
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        if not rows:
            return []

        # Criar lista de (nome_normalizado, cliente_original)
        clientes = [(normalizar_nome(r.get("nome", "")), dict(r)) for r in rows if r.get("nome")]

        if not clientes:
            return []

        # Usar rapidfuzz para encontrar top matches
        nomes_normalizados = [c[0] for c in clientes]
        resultados = process.extract(
            nome_normalizado,
            nomes_normalizados,
            scorer=fuzz.token_sort_ratio,
            limit=limite
        )

        # Montar resultado com scores
        matches = []
        for nome_match, score, idx in resultados:
            if score >= 30:  # Mínimo 30% para aparecer como sugestão
                cliente = clientes[idx][1]
                matches.append({
                    **cliente,
                    "score": score / 100,  # Converter para 0-1
                    "score_percent": score
                })

        return matches
    except Exception as e:
        logger.error(f"Erro ao buscar clientes similares: {e}")
        return []


def verificar_andamento_existente(processo_id: int, data: str, descricao: str) -> bool:
    """Verifica se um andamento já existe (para evitar duplicatas)."""
    conn = get_db()
    if not conn:
        return False

    try:
        cur = conn.cursor()
        # Verifica por data + início da descrição (para evitar problemas com texto truncado)
        cur.execute("""
            SELECT 1 FROM processo_andamentos
            WHERE processo_id = %s AND data = %s AND descricao ILIKE %s
            LIMIT 1
        """, (processo_id, data, descricao[:100] + "%"))
        existe = cur.fetchone() is not None
        cur.close()
        conn.close()
        return existe
    except Exception as e:
        logger.error(f"Erro ao verificar andamento existente: {e}")
        return False


@app.post("/api/admin/importar-astrea/preview")
async def preview_importacao_astrea(
    arquivo: UploadFile = File(...),
    admin = Depends(verificar_admin)
):
    """
    Faz preview dos dados do arquivo Excel antes de importar.
    Retorna lista de processos e andamentos encontrados.
    """
    if not arquivo.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser Excel (.xlsx ou .xls)")

    try:
        conteudo = await arquivo.read()
        resultado = parsear_excel_astrea(conteudo)

        if "erro" in resultado:
            raise HTTPException(status_code=400, detail=resultado["erro"])

        # Enriquecer dados com informações do sistema
        for processo in resultado["processos"]:
            # Verificar se processo já existe
            processo_existente = buscar_processo_por_numero(processo["numero_processo"])
            processo["existe_no_sistema"] = processo_existente is not None
            if processo_existente:
                processo["processo_id_existente"] = processo_existente["id"]
                processo["cadastro_id_existente"] = processo_existente["cadastro_id"]

            # Tentar vincular cliente pelo CPF ou Nome (APENAS CORRESPONDÊNCIA EXATA)
            cliente = None
            processo["cliente_encontrado"] = False

            # 1. Primeiro tentar por CPF (mais preciso)
            if processo.get("cliente_cpf"):
                cliente = buscar_cliente_por_cpf(processo["cliente_cpf"])
                if cliente:
                    processo["match_tipo"] = "cpf"

            # 2. Se não encontrou por CPF, tentar por nome com FUZZY MATCHING
            if not cliente and processo.get("cliente_nome"):
                # Primeiro tenta exato
                cliente = buscar_cliente_por_nome(processo["cliente_nome"])
                if cliente:
                    processo["match_tipo"] = "nome_exato"
                    processo["match_score"] = 100
                else:
                    # Tentar fuzzy matching com threshold de 90% (mais rigoroso)
                    cliente = buscar_cliente_por_nome_fuzzy(processo["cliente_nome"], threshold=90)
                    if cliente:
                        processo["match_tipo"] = "nome_fuzzy"
                        processo["match_score"] = cliente.get("match_score", 0)
                        logger.info(f"Fuzzy match: '{processo['cliente_nome']}' -> '{cliente.get('nome')}' (score: {processo['match_score']})")

            if cliente:
                processo["cliente_encontrado"] = True
                processo["cadastro_id_sugerido"] = cliente["id"]
                processo["cliente_nome_sistema"] = cliente.get("nome", "")

            # Contar andamentos novos
            if processo_existente:
                novos_andamentos = 0
                for and_item in processo.get("andamentos", []):
                    if not verificar_andamento_existente(
                        processo_existente["id"],
                        and_item["data"],
                        and_item["descricao"]
                    ):
                        novos_andamentos += 1
                processo["andamentos_novos"] = novos_andamentos
            else:
                processo["andamentos_novos"] = len(processo.get("andamentos", []))

        # Adicionar estatísticas de correspondência
        clientes_encontrados = sum(1 for p in resultado["processos"] if p.get("cliente_encontrado"))
        clientes_nao_encontrados = len(resultado["processos"]) - clientes_encontrados

        resultado["estatisticas"] = {
            "total_processos": len(resultado["processos"]),
            "clientes_encontrados": clientes_encontrados,
            "clientes_nao_encontrados": clientes_nao_encontrados,
            "percentual_match": round(clientes_encontrados / max(len(resultado["processos"]), 1) * 100, 1)
        }

        # Adicionar aviso se muitos clientes não foram encontrados
        if clientes_nao_encontrados > 0:
            resultado["erros"] = resultado.get("erros", [])
            resultado["erros"].append(
                f"{clientes_nao_encontrados} processo(s) sem cliente correspondente no sistema. "
                "Selecione um cliente padrão ou cadastre os clientes primeiro."
            )

        return resultado

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro no preview de importação: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class ImportacaoAstreaRequest(BaseModel):
    processos: List[dict]
    cadastro_id_padrao: Optional[str] = None  # Para vincular todos ao mesmo cliente
    cliente_id: Optional[str] = None  # Alias para cadastro_id_padrao (frontend usa este)
    importar_andamentos: bool = True
    andamentos_visiveis: bool = True
    arquivo_nome: Optional[str] = None
    arquivo_tamanho: Optional[int] = None


@app.post("/api/admin/importar-astrea/confirmar")
async def confirmar_importacao_astrea(
    dados: ImportacaoAstreaRequest,
    admin = Depends(verificar_admin)
):
    """
    Efetua a importação dos processos e andamentos do Astrea.
    """
    # Suporte a cliente_id como alias de cadastro_id_padrao
    cadastro_id_padrao = dados.cadastro_id_padrao or dados.cliente_id

    relatorio = {
        "processos_criados": 0,
        "processos_atualizados": 0,
        "andamentos_criados": 0,
        "erros": [],
        "detalhes": []
    }

    for processo_data in dados.processos:
        try:
            numero_processo = processo_data.get("numero_processo")
            if not numero_processo:
                relatorio["erros"].append("Processo sem número ignorado")
                continue

            # Determinar cadastro_id
            cadastro_id = None

            # 1. Se tem cadastro_id padrão, usar
            if cadastro_id_padrao:
                cadastro_id = cadastro_id_padrao
            # 2. Se processo já existe, manter vínculo existente
            elif processo_data.get("cadastro_id_existente"):
                cadastro_id = processo_data["cadastro_id_existente"]
            # 3. Se encontrou cliente pelo CPF
            elif processo_data.get("cadastro_id_sugerido"):
                cadastro_id = processo_data["cadastro_id_sugerido"]

            if not cadastro_id:
                relatorio["erros"].append(f"Processo {numero_processo}: Cliente não identificado")
                continue

            # Verificar se processo existe
            processo_existente = buscar_processo_por_numero(numero_processo)

            processo_dados = {
                "numero_processo": numero_processo,
                "tipo_acao": processo_data.get("tipo_acao", ""),
                "vara_tribunal": processo_data.get("vara_tribunal", ""),
                "fase": processo_data.get("fase", "Inicial"),
                "reu": processo_data.get("reu", ""),
                "valor_causa": processo_data.get("valor_causa", 0),
                "data_distribuicao": processo_data.get("data_distribuicao"),
                "observacoes": processo_data.get("observacoes", ""),
                "status": "ativo"
            }

            if processo_existente:
                # Atualizar processo existente
                processo_id = processo_existente["id"]
                if atualizar_processo(processo_id, processo_dados):
                    relatorio["processos_atualizados"] += 1
                    relatorio["detalhes"].append({
                        "tipo": "atualizado",
                        "numero": numero_processo,
                        "processo_id": processo_id
                    })
                else:
                    relatorio["erros"].append(f"Erro ao atualizar processo {numero_processo}")
                    continue
            else:
                # Criar novo processo
                processo_id = criar_processo(cadastro_id, processo_dados)
                if processo_id:
                    relatorio["processos_criados"] += 1
                    relatorio["detalhes"].append({
                        "tipo": "criado",
                        "numero": numero_processo,
                        "processo_id": processo_id,
                        "cadastro_id": cadastro_id
                    })
                else:
                    relatorio["erros"].append(f"Erro ao criar processo {numero_processo}")
                    continue

            # Importar andamentos
            if dados.importar_andamentos and processo_data.get("andamentos"):
                for andamento in processo_data["andamentos"]:
                    # Verificar se andamento já existe
                    if verificar_andamento_existente(
                        processo_id,
                        andamento["data"],
                        andamento["descricao"]
                    ):
                        continue

                    # Criar andamento
                    and_id = criar_andamento_processo(
                        processo_id,
                        andamento["data"],
                        andamento["descricao"],
                        dados.andamentos_visiveis
                    )
                    if and_id:
                        relatorio["andamentos_criados"] += 1
                    else:
                        relatorio["erros"].append(
                            f"Erro ao criar andamento para processo {numero_processo}"
                        )

        except Exception as e:
            logger.error(f"Erro ao importar processo: {e}")
            relatorio["erros"].append(f"Erro inesperado: {str(e)}")

    relatorio["sucesso"] = True
    relatorio["total_erros"] = len(relatorio["erros"])

    # Salvar histórico da importação
    try:
        conn = get_db()
        if conn:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO historico_importacoes
                (tipo, arquivo_nome, arquivo_tamanho, processos_criados, processos_atualizados,
                 andamentos_adicionados, erros, detalhes, usuario_id, usuario_nome)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                'astrea',
                dados.arquivo_nome or 'arquivo.xlsx',
                dados.arquivo_tamanho or 0,
                relatorio["processos_criados"],
                relatorio["processos_atualizados"],
                relatorio["andamentos_criados"],
                len(relatorio["erros"]),
                json.dumps({"erros": relatorio["erros"][:10], "detalhes": relatorio["detalhes"][:20]}),
                admin.get("id"),
                admin.get("nome", "Admin")
            ))
            conn.commit()
            cur.close()
            conn.close()
    except Exception as e:
        logger.error(f"Erro ao salvar histórico de importação: {e}")

    return relatorio


# ============================================
# INTEGRAÇÃO DATAJUD - MOVIDO PARA routes/datajud.py
# ============================================

# ============================================
# PRAZOS PROCESSUAIS - MOVIDO PARA routes/prazos.py
# Endpoints: /api/admin/prazos/*, /api/admin/processos/*/prazos, /api/admin/clientes/*/prazos
# ============================================

# ============================================
# ENDPOINT DE IA PARA EXPLICAR ANDAMENTOS
# ============================================

@app.post("/api/cliente/andamento/explicar")
async def explicar_andamento_ia(
    dados: dict,
    authorization: str = Header(None)
):
    """
    Usa IA para explicar tecnicamente um andamento processual.
    Retorna explicação em linguagem acessível.
    """
    # Verificar token do cliente
    if not authorization:
        raise HTTPException(status_code=401, detail="Token não fornecido")

    token = authorization.replace("Bearer ", "")
    cliente = decodificar_token_cliente(token)
    if not cliente:
        raise HTTPException(status_code=401, detail="Token inválido")

    descricao = dados.get("descricao", "")
    if not descricao:
        raise HTTPException(status_code=400, detail="Descrição do andamento não fornecida")

    # Criar explicação usando template (pode ser substituído por API de IA como Claude)
    explicacao = gerar_explicacao_andamento(descricao)

    return {
        "sucesso": True,
        "andamento_original": descricao,
        "explicacao": explicacao,
        "aviso": "Esta explicação foi gerada automaticamente e pode conter imprecisões. Para esclarecimentos, utilize o módulo de mensagens."
    }


def gerar_explicacao_andamento(descricao: str) -> str:
    """
    Gera uma explicação simplificada do andamento processual.
    Pode ser integrado com API de IA (Claude, GPT) para explicações mais sofisticadas.
    """
    descricao_lower = descricao.lower()

    # Dicionário de termos jurídicos comuns e suas explicações
    explicacoes = {
        "distribuído": "O processo foi registrado e encaminhado para uma vara/juízo específico que será responsável por analisá-lo.",
        "distribuição": "O processo foi registrado e encaminhado para uma vara/juízo específico que será responsável por analisá-lo.",
        "citação": "O réu (parte contrária) está sendo oficialmente notificado sobre a existência do processo e terá prazo para se defender.",
        "citado": "O réu (parte contrária) foi oficialmente notificado sobre o processo.",
        "contestação": "A parte contrária apresentou sua defesa, respondendo às acusações ou pedidos feitos.",
        "sentença": "O juiz proferiu sua decisão sobre o caso. Esta decisão pode ser favorável ou desfavorável e ainda pode ser objeto de recurso.",
        "julgado": "O caso foi analisado e decidido pelo juiz ou tribunal.",
        "audiência": "Foi marcada ou realizada uma sessão presencial ou virtual onde as partes e o juiz discutem o caso.",
        "despacho": "O juiz emitiu uma ordem ou determinação sobre algum aspecto do processo.",
        "intimação": "Uma das partes está sendo notificada sobre alguma decisão ou prazo no processo.",
        "intimado": "A parte foi notificada oficialmente sobre uma decisão ou prazo.",
        "recurso": "Uma das partes está contestando uma decisão anterior, pedindo que seja revista por instância superior.",
        "apelação": "Recurso apresentado contra a sentença, pedindo que o tribunal de segunda instância revise a decisão.",
        "agravo": "Recurso contra decisões intermediárias do processo, geralmente sobre questões processuais.",
        "embargos": "Recurso que pede esclarecimento, correção ou complementação de uma decisão.",
        "trânsito em julgado": "A decisão se tornou definitiva, não cabendo mais recursos. O processo está encerrado nesta fase.",
        "arquivado": "O processo foi encerrado e guardado, seja por decisão final ou por outros motivos legais.",
        "baixa": "O processo foi devolvido à instância inferior ou encerrado.",
        "petição": "Um documento foi apresentado por uma das partes solicitando algo ao juiz.",
        "concluso": "O processo está com o juiz para análise e decisão.",
        "vista": "O processo foi encaminhado para uma das partes ou para o Ministério Público analisar.",
        "perícia": "Foi determinada ou realizada uma análise técnica por especialista sobre algum aspecto do caso.",
        "penhora": "Bens foram bloqueados ou apreendidos para garantir o pagamento de uma dívida.",
        "execução": "Fase em que a decisão judicial está sendo cumprida, geralmente envolvendo cobrança de valores.",
        "cumprimento": "A parte está executando (cumprindo) o que foi determinado na sentença.",
        "acordo": "As partes chegaram a um entendimento e resolveram a questão de forma consensual.",
        "homologação": "O juiz aprovou oficialmente um acordo ou documento apresentado pelas partes.",
        "tutela": "Medida de proteção ou antecipação de direitos concedida pelo juiz.",
        "liminar": "Decisão urgente tomada no início do processo para proteger direitos enquanto o caso é analisado.",
        "suspensão": "O processo foi temporariamente paralisado por algum motivo legal.",
        "expedição": "Um documento oficial (como mandado ou ofício) foi emitido pelo cartório.",
        "certidão": "Documento que atesta ou comprova algum fato ou situação do processo.",
        "autos": "Conjunto de todos os documentos e peças do processo.",
        "remessa": "O processo foi enviado para outro órgão, instância ou localidade.",
        "juntada": "Um documento foi anexado ao processo.",
    }

    # Buscar explicação correspondente
    for termo, explicacao in explicacoes.items():
        if termo in descricao_lower:
            return f"📋 **Explicação:** {explicacao}\n\n💡 **Contexto:** Este andamento indica uma movimentação importante no seu processo. O andamento '{descricao}' significa que houve uma ação relacionada a '{termo}'."

    # Explicação genérica se não encontrar termo específico
    return f"📋 **Explicação:** Este é um andamento processual que indica uma movimentação no seu processo. Cada andamento representa uma etapa ou ação realizada pelo judiciário.\n\n💡 **O que significa '{descricao}':** Este registro indica que houve uma atividade no seu processo. Para entender melhor o impacto desta movimentação no seu caso específico, recomendamos entrar em contato através do módulo de mensagens."


@app.get("/api/admin/importacoes/historico")
async def listar_historico_importacoes(admin = Depends(verificar_admin)):
    """Lista o histórico de importações."""
    try:
        conn = get_db()
        if not conn:
            raise HTTPException(status_code=500, detail="Erro de conexão")

        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, tipo, arquivo_nome, arquivo_tamanho,
                   processos_criados, processos_atualizados, andamentos_adicionados,
                   erros, detalhes, usuario_nome, criado_em
            FROM historico_importacoes
            ORDER BY criado_em DESC
            LIMIT 50
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        historico = []
        for row in rows:
            historico.append({
                "id": row["id"],
                "tipo": row["tipo"],
                "arquivo_nome": row["arquivo_nome"],
                "arquivo_tamanho": row["arquivo_tamanho"],
                "processos_criados": row["processos_criados"],
                "processos_atualizados": row["processos_atualizados"],
                "andamentos_adicionados": row["andamentos_adicionados"],
                "erros": row["erros"],
                "detalhes": row["detalhes"] if isinstance(row["detalhes"], dict) else json.loads(row["detalhes"] or "{}"),
                "usuario_nome": row["usuario_nome"],
                "criado_em": row["criado_em"].isoformat() if row["criado_em"] else None
            })

        return historico

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao listar histórico: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/debug/clientes-cadastrados")
async def debug_clientes():
    """TEMPORÁRIO: Ver clientes cadastrados."""
    try:
        conn = get_db()
        if not conn:
            return {"erro": "Sem conexão"}
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT id, dados->>'nome' as nome, dados->>'cpf' as cpf FROM cadastros LIMIT 50")
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return {"clientes": [dict(r) for r in rows]}
    except Exception as e:
        return {"erro": str(e)}


@app.get("/api/debug/testar-match/{nome}")
async def debug_testar_match(nome: str):
    """TEMPORÁRIO: Testar matching de nome."""
    try:
        nome_normalizado = normalizar_nome(nome)
        cliente_exato = buscar_cliente_por_nome_fuzzy(nome, threshold=75)
        similares = buscar_clientes_similares(nome, 5)
        return {
            "nome_original": nome,
            "nome_normalizado": nome_normalizado,
            "match_encontrado": cliente_exato,
            "similares": similares
        }
    except Exception as e:
        return {"erro": str(e)}


@app.post("/api/debug/analisar-excel-clientes")
async def debug_analisar_excel_clientes(arquivo: UploadFile = File(...)):
    """
    DEBUG: Analisa o Excel e mostra todos os nomes únicos de clientes encontrados,
    junto com os resultados de matching.
    """
    try:
        conteudo = await arquivo.read()
        resultado = parsear_excel_astrea(conteudo)

        if "erro" in resultado:
            return {"erro": resultado["erro"]}

        # Coletar todos os nomes únicos de clientes
        nomes_unicos = set()
        for processo in resultado.get("processos", []):
            nome = processo.get("cliente_nome", "")
            if nome:
                nomes_unicos.add(nome)

        # Para cada nome, verificar se encontra correspondência
        analise = []
        for nome in sorted(nomes_unicos):
            nome_normalizado = normalizar_nome(nome)

            # Buscar match fuzzy
            cliente_match = buscar_cliente_por_nome_fuzzy(nome, threshold=75)

            # Buscar similares (mesmo sem threshold)
            similares = buscar_clientes_similares(nome, 3)

            analise.append({
                "nome_excel": nome,
                "nome_normalizado": nome_normalizado,
                "encontrou_match": cliente_match is not None,
                "match": {
                    "id": cliente_match.get("id") if cliente_match else None,
                    "nome": cliente_match.get("nome") if cliente_match else None,
                    "score": cliente_match.get("match_score") if cliente_match else None
                } if cliente_match else None,
                "similares": similares[:3] if similares else []
            })

        # Contar matches
        total_nomes = len(nomes_unicos)
        total_matches = sum(1 for a in analise if a["encontrou_match"])

        return {
            "total_processos": len(resultado.get("processos", [])),
            "total_nomes_unicos": total_nomes,
            "total_matches": total_matches,
            "percentual_match": round(total_matches / max(total_nomes, 1) * 100, 1),
            "headers_excel": resultado.get("headers_encontrados", [])[:15],
            "coluna_cliente_detectada": "cliente_nome" in resultado.get("colunas_encontradas", []),
            "analise_detalhada": analise
        }
    except Exception as e:
        logger.error(f"Erro ao analisar Excel: {e}")
        return {"erro": str(e)}


@app.get("/api/admin/processos/limpar-todos-agora")
async def limpar_todos_processos_get(admin = Depends(verificar_admin)):
    """Versão GET para limpar todos os processos (mais compatível)."""
    return await limpar_todos_processos_impl(admin)


@app.delete("/api/admin/processos/limpar-todos")
async def limpar_todos_processos(admin = Depends(verificar_admin)):
    """Versão DELETE para limpar todos os processos."""
    return await limpar_todos_processos_impl(admin)


async def limpar_todos_processos_impl(admin):
    """
    CUIDADO: Deleta TODOS os processos e andamentos do sistema.
    Usar apenas para limpar importações incorretas.
    """
    try:
        conn = get_db()
        if not conn:
            raise HTTPException(status_code=500, detail="Erro de conexão")

        cur = conn.cursor()

        # Contar antes de deletar
        cur.execute("SELECT COUNT(*) FROM processos")
        total_processos = cur.fetchone()[0]

        cur.execute("SELECT COUNT(*) FROM processo_andamentos")
        total_andamentos = cur.fetchone()[0]

        # Deletar andamentos primeiro (FK constraint)
        cur.execute("DELETE FROM processo_andamentos")

        # Deletar processos
        cur.execute("DELETE FROM processos")

        conn.commit()
        cur.close()
        conn.close()

        logger.info(f"Admin {admin.get('nome')} deletou {total_processos} processos e {total_andamentos} andamentos")

        return {
            "sucesso": True,
            "processos_deletados": total_processos,
            "andamentos_deletados": total_andamentos,
            "mensagem": f"Deletados {total_processos} processos e {total_andamentos} andamentos"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao limpar processos: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/clientes-para-importacao")
async def listar_clientes_para_importacao(admin = Depends(verificar_admin)):
    """Lista clientes para seleção na importação do Astrea."""
    try:
        cadastros = carregar_cadastros()
        clientes = [
            {
                "id": c["id"],
                "nome": c["dados"].get("nome", ""),
                "cpf": c["dados"].get("cpf", ""),
                "email": c["dados"].get("email", "")
            }
            for c in cadastros
        ]
        return {"clientes": clientes}
    except Exception as e:
        logger.error(f"Erro ao listar clientes: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# BANNERS E NOTÍCIAS - MOVIDO PARA routes/banners.py
# Endpoints: /api/admin/banners/*, /api/cliente/banners
# ============================================

# ============================================
# INICIALIZAÇÃO
# ============================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
